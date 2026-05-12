"""
Тесты слоя io_tools.
"""
from datetime import date, datetime, time
import logging
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from app.io_tools import (
    DataNormalizer,
    DataProcessingResult,
    DataProcessor,
    DataResolver,
    ExtractedTable,
    HeaderBinding,
    ImportExportService,
    ImportProcessor,
    RawWorkbookReader,
    ProcessedRow,
    StrictImportProcessor,
    TableRegion,
    XlsxExporter,
    XlsxImporter,
    XlsxRangeReader,
)
from app.io_tools.xlsx_config import (
    SMART_IMPORT_ENTITY_TYPES,
    STRICT_IMPORT_ENTITY_TYPES,
    XLSX_COLUMNS_BY_SHEET,
    XLSX_SHEETS_ORDER,
)
from app.models import (
    Attendance,
    Comment,
    Group,
    Lesson,
    Mark,
    Schedule,
    ScheduleGroupLink,
    Student,
)
from app.services import OrmService


class SampleSchema(BaseModel):
    """
    Вспомогательная схема для тестов io_tools.
    """

    id: int
    title: str


def _logs_dir() -> Path:
    """
    Возвращает путь к директории логов проекта.

    :return: Абсолютный путь к папке `logs` в корне проекта.
    """
    return Path(__file__).resolve().parents[1] / "logs"


def _normalize_db_row(row: object, sheet_name: str) -> dict[str, object]:
    """
    Преобразует ORM-объект в словарь значений по конфигурации XLSX-листа.

    :param row: ORM-объект, представляющий строку таблицы.
    :param sheet_name: Каноническое имя XLSX-листа для этой модели.
    :return: Словарь значений в порядке и составе, используемом для экспорта.
    """
    return {
        column_name: getattr(row, column_name)
        for column_name in XLSX_COLUMNS_BY_SHEET[sheet_name]
    }


def _sorted_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    """
    Сортирует строки листа по идентификатору для стабильного сравнения.

    :param rows: Список строк одного листа.
    :return: Новый список строк, отсортированный по полю `id`.
    """
    return sorted(rows, key=lambda row: row["id"])


def _normalize_imported_rows(
    imported_rows: list[dict[str, object]],
    expected_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    """
    Приводит импортированные из XLSX значения к виду, пригодному для точного сравнения.

    Excel иногда возвращает дату как `datetime` с нулевым временем. Для round-trip
    проверки это не считается искажением данных, поэтому функция нормализует такие
    значения по типу соответствующего поля из ожидаемой строки.

    :param imported_rows: Строки, считанные из XLSX-файла.
    :param expected_rows: Эталонные строки, собранные из ORM-объектов БД.
    :return: Новый список импортированных строк после нормализации типов значений.
    """
    normalized_rows: list[dict[str, object]] = []

    for imported_row, expected_row in zip(
        _sorted_rows(imported_rows),
        _sorted_rows(expected_rows),
        strict=True,
    ):
        normalized_row: dict[str, object] = {}

        for key, imported_value in imported_row.items():
            expected_value = expected_row[key]
            if (
                isinstance(imported_value, datetime)
                and isinstance(expected_value, date)
                and not isinstance(expected_value, datetime)
            ):
                normalized_row[key] = imported_value.date()
            else:
                normalized_row[key] = imported_value

        normalized_rows.append(normalized_row)

    return normalized_rows


def test_xlsx_exporter_exports_full_and_partial_payloads(tmp_path):
    """
    Один и тот же метод экспортирует как полный, так и частичный набор листов.
    """
    exporter = XlsxExporter()
    full_path = tmp_path / "full_export.xlsx"
    partial_path = tmp_path / "partial_export.xlsx"

    full_payload = {
        "groups": [
            {
                "id": 1,
                "name": "IU7-11",
                "speciality": "09.03.01_Informatics",
                "created_at": datetime(2026, 5, 10, 12, 30, 0),
            },
            Group(name="IU7-12", speciality="09.03.02_Applied"),
        ],
        "lessons": [
            {
                "id": 3,
                "date": date(2026, 5, 10),
                "time": time(9, 0),
                "topic": "SQLAlchemy",
            }
        ],
    }
    partial_payload = {
        "groups": full_payload["groups"],
    }

    exporter.export(full_payload, full_path)
    exporter.export(partial_payload, partial_path)

    full_workbook = load_workbook(full_path)
    partial_workbook = load_workbook(partial_path)

    assert full_workbook.sheetnames == ["groups", "lessons"]
    assert partial_workbook.sheetnames == ["groups"]

    groups_sheet = full_workbook["groups"]
    assert groups_sheet.cell(1, 1).value == "id"
    assert groups_sheet.cell(1, 2).value == "name"
    assert groups_sheet.cell(2, 2).value == "IU7-11"
    assert groups_sheet.cell(3, 2).value == "IU7-12"

    lessons_sheet = full_workbook["lessons"]
    assert lessons_sheet.cell(1, 2).value == "schedule_id"
    assert lessons_sheet.cell(1, 3).value == "topic"
    assert lessons_sheet.cell(1, 4).value == "date"
    assert lessons_sheet.cell(2, 3).value == "SQLAlchemy"


def test_xlsx_exporter_uses_global_sheet_order(tmp_path):
    """
    При полном или смешанном экспорте известные листы идут в порядке конфигурации.
    """
    exporter = XlsxExporter()
    file_path = tmp_path / "ordered_export.xlsx"
    payload = {
        "students": [{"id": 1, "surname": "Petrov", "first_name": "Petr"}],
        "groups": [{"id": 1, "name": "IU7-11"}],
        "custom": [{"value": "note"}],
        "comments": [{"id": 1, "student_id": 1, "lesson_id": 1}],
    }

    exporter.export(payload, file_path)
    workbook = load_workbook(file_path)

    assert workbook.sheetnames == ["groups", "students", "comments", "custom"]


def test_xlsx_exporter_supports_schema_rows(tmp_path):
    """
    Экспортер принимает строки в виде схем и обычных словарей.
    """
    exporter = XlsxExporter()
    file_path = tmp_path / "schema_export.xlsx"
    payload = {
        "meta": [
            SampleSchema(id=1, title="Header"),
            {"id": 2, "title": "Body"},
        ]
    }

    saved_path = exporter.export(payload, file_path)
    workbook = load_workbook(saved_path)
    meta_sheet = workbook["meta"]

    assert isinstance(saved_path, Path)
    assert meta_sheet.cell(1, 1).value == "id"
    assert meta_sheet.cell(1, 2).value == "title"
    assert meta_sheet.cell(2, 2).value == "Header"
    assert meta_sheet.cell(3, 2).value == "Body"


def test_xlsx_exporter_rejects_empty_payload(tmp_path):
    """
    Пустой экспортный набор считается ошибкой.
    """
    exporter = XlsxExporter()

    with pytest.raises(
        ValueError,
        match="Для экспорта нужно передать хотя бы один лист.",
    ):
        exporter.export({}, tmp_path / "empty.xlsx")


def test_xlsx_importer_reads_known_sheets_in_configured_order(tmp_path):
    """
    Импортер читает листы в порядке XLSX-конфигурации и пропускает пустые строки.
    """
    exporter = XlsxExporter()
    importer = XlsxImporter()
    file_path = tmp_path / "importable_export.xlsx"
    payload = {
        "students": [
            {"id": 1, "group_id": 1, "surname": "Petrov", "first_name": "Petr"},
            {"id": None, "group_id": None, "surname": None, "first_name": None},
        ],
        "groups": [{"id": 1, "name": "IU7-11", "speciality": "09.03.01"}],
    }

    exporter.export(payload, file_path)
    imported = importer.import_data(file_path)

    assert list(imported) == ["groups", "students"]
    assert imported["groups"][0]["name"] == "IU7-11"
    assert imported["students"][0]["surname"] == "Petrov"
    assert len(imported["students"]) == 1


def test_xlsx_importer_rejects_unknown_sheet(tmp_path):
    """
    Импортер отклоняет файл с неизвестным листом.
    """
    exporter = XlsxExporter()
    importer = XlsxImporter()
    file_path = tmp_path / "unknown_sheet.xlsx"
    payload = {
        "custom": [{"value": "note"}],
    }

    exporter.export(payload, file_path)

    with pytest.raises(
        ValueError,
        match="Неизвестные листы в XLSX-файле: custom.",
    ):
        importer.import_data(file_path)


def test_xlsx_importer_rejects_missing_required_columns(tmp_path):
    """
    Импортер проверяет наличие обязательных колонок.
    """
    exporter = XlsxExporter()
    importer = XlsxImporter()
    file_path = tmp_path / "missing_headers.xlsx"
    payload = {
        "students": [{"id": 1, "surname": "Petrov"}],
    }

    exporter.export(payload, file_path)
    workbook = load_workbook(file_path)
    worksheet = workbook["students"]
    worksheet.delete_cols(2, 3)
    workbook.save(file_path)

    with pytest.raises(
        ValueError,
        match="В листе students отсутствуют обязательные колонки: group_id, surname, first_name.",
    ):
        importer.import_data(file_path)


def test_import_export_service_delegates_export_and_import(tmp_path):
    """
    Верхнеуровневый сервис делегирует экспорт и импорт XLSX соответствующим обработчикам.
    """
    service = ImportExportService()
    file_path = tmp_path / "service_import.xlsx"
    payload = {
        "groups": [{"id": 1, "name": "IU7-11"}],
    }

    exported_path = service.export_to_xlsx(payload, file_path)
    imported = service.import_from_xlsx(file_path)

    assert exported_path == file_path
    assert imported["groups"][0]["name"] == "IU7-11"


def test_xlsx_importer_finds_table_candidates_outside_a1(tmp_path):
    """
    Поиск нестандартных таблиц находит таблицу, даже если она начинается не в A1.
    """
    importer = XlsxImporter()
    file_path = tmp_path / "offset_table.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Students"
    worksheet["A1"] = "Список студентов"
    worksheet["C4"] = "surname"
    worksheet["D4"] = "first_name"
    worksheet["E4"] = "group_id"
    worksheet["C5"] = "Иванов"
    worksheet["D5"] = "Иван"
    worksheet["E5"] = 1
    worksheet["C6"] = "Петров"
    worksheet["D6"] = "Петр"
    worksheet["E6"] = 1
    workbook.save(file_path)

    tables = importer.find_table_candidates(file_path)

    assert len(tables) == 1
    assert tables[0].sheet == "Students"
    assert tables[0].range == "C4:E6"
    assert tables[0].rows == 3
    assert tables[0].cols == 3


def test_xlsx_importer_reads_selected_range_and_keeps_extra_columns(tmp_path):
    """
    Чтение диапазона сохраняет лишние колонки и отдельно сообщает о них.
    """
    importer = XlsxImporter()
    file_path = tmp_path / "selected_range.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["B3"] = "surname"
    worksheet["C3"] = "first_name"
    worksheet["D3"] = "group_id"
    worksheet["E3"] = "Комментарий"
    worksheet["B4"] = "Иванов"
    worksheet["C4"] = "Иван"
    worksheet["D4"] = 1
    worksheet["E4"] = "Староста"
    workbook.save(file_path)

    table = importer.read_table_range(
        file_path,
        "Sheet1",
        "B3:E4",
        entity_type="student",
    )

    assert table.entity_type == "students"
    assert table.headers == ("surname", "first_name", "group_id", "Комментарий")
    assert table.known_headers == ("surname", "first_name", "group_id")
    assert table.unknown_headers == ("Комментарий",)
    assert table.missing_required_headers == ()
    assert table.rows == [
        {
            "surname": "Иванов",
            "first_name": "Иван",
            "group_id": 1,
            "Комментарий": "Староста",
        }
    ]
    assert table.errors == []
    assert table.is_valid is True


def test_xlsx_importer_reports_missing_required_headers_for_selected_range(tmp_path):
    """
    Чтение диапазона сообщает о пропущенных обязательных колонках выбранной сущности.
    """
    importer = XlsxImporter()
    file_path = tmp_path / "bad_range.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["D5"] = "surname"
    worksheet["E5"] = "Комментарий"
    worksheet["D6"] = "Иванов"
    worksheet["E6"] = "Нет имени"
    workbook.save(file_path)

    table = importer.read_table_range(
        file_path,
        "Sheet1",
        "D5:E6",
        entity_type="student",
    )

    assert table.entity_type == "students"
    assert table.missing_required_headers == ("group_id", "first_name")
    assert table.errors == ["Missing required headers: group_id, first_name."]
    assert table.is_valid is False


def test_raw_workbook_reader_normalizes_empty_like_values():
    """
    Низкоуровневый reader одинаково трактует пустые значения разных видов.
    """
    assert RawWorkbookReader.normalize_cell_value(None) is None
    assert RawWorkbookReader.normalize_cell_value("   ") is None
    assert RawWorkbookReader.normalize_cell_value(" text ") == "text"
    assert RawWorkbookReader.normalize_cell_value(0) == 0
    assert RawWorkbookReader.is_non_empty(" value ") is True
    assert RawWorkbookReader.is_non_empty("") is False


def test_xlsx_range_reader_reports_generated_and_duplicated_headers(tmp_path):
    """
    Reader сообщает о пустых и дублирующихся заголовках, а строки при этом не теряются.
    """
    reader = XlsxRangeReader()
    file_path = tmp_path / "duplicate_headers.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["C3"] = "surname"
    worksheet["D3"] = None
    worksheet["E3"] = "surname"
    worksheet["C4"] = "Иванов"
    worksheet["D4"] = "заметка"
    worksheet["E4"] = "Петров"
    workbook.save(file_path)

    table = reader.read_range(file_path, "Sheet1", "C3:E4")

    assert table.headers == ("surname", "__column_2", "surname__3")
    assert table.rows == [
        {
            "surname": "Иванов",
            "__column_2": "заметка",
            "surname__3": "Петров",
        }
    ]
    assert table.warnings == [
        "Empty header in column 2 was replaced with __column_2.",
        "Duplicate header was renamed to surname__3.",
        "Header surname appeared more than once in the selected range.",
    ]
    assert table.errors == []


def test_xlsx_range_reader_reports_missing_data_rows(tmp_path):
    """
    Если в выбранном диапазоне есть только заголовки, reader возвращает явную ошибку.
    """
    reader = XlsxRangeReader()
    file_path = tmp_path / "headers_only.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["B2"] = "surname"
    worksheet["C2"] = "first_name"
    worksheet["D2"] = "group_id"
    workbook.save(file_path)

    table = reader.read_range(
        file_path,
        "Sheet1",
        "B2:D2",
        entity_type="student",
    )

    assert table.headers == ("surname", "first_name", "group_id")
    assert table.rows == []
    assert table.errors == ["Selected range does not contain data rows."]
    assert table.is_valid is False


def test_xlsx_importer_reads_detected_table_region(tmp_path):
    """
    Импортер умеет читать уже найденный диапазон через объект TableRegion.
    """
    importer = XlsxImporter()
    file_path = tmp_path / "detected_table.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["F6"] = "surname"
    worksheet["G6"] = "first_name"
    worksheet["H6"] = "group_id"
    worksheet["F7"] = "Иванов"
    worksheet["G7"] = "Иван"
    worksheet["H7"] = 3
    workbook.save(file_path)

    table_region = TableRegion(
        sheet="Sheet1",
        range="F6:H7",
        min_row=6,
        max_row=7,
        min_col=6,
        max_col=8,
        rows=2,
        cols=3,
        total_cells=6,
        non_empty_cells=6,
        density=1.0,
        score=0.9,
    )

    table = importer.read_detected_table(
        file_path,
        table_region,
        entity_type="student",
    )

    assert table.entity_type == "students"
    assert table.rows == [
        {"surname": "Иванов", "first_name": "Иван", "group_id": 3}
    ]
    assert table.errors == []


def test_smart_reader_detects_and_reads_real_nonstandard_tables():
    """
    Нестандартный XLSX-файл из `logs` корректно разбирается на три таблицы.

    Тест проверяет реальный сценарий умного чтения:
    - в книге есть мусор вокруг таблиц;
    - таблицы начинаются не из `A1`;
    - заголовки не совпадают с внутренним форматом приложения;
    - сами строки при этом все равно должны быть прочитаны без потери данных.
    """
    service = ImportExportService()
    file_path = _logs_dir() / "test_io_tools_smart_reader.xlsx"
    if not file_path.exists():
        pytest.skip(f"Тестовый XLSX-файл не найден: {file_path}")

    tables = service.find_xlsx_tables(file_path)

    assert [(table.sheet, table.range) for table in tables] == [
        ("Лист1", "C3:D10"),
        ("Лист1", "L12:O16"),
        ("Лист2", "C5:D11"),
    ]

    first_groups = service.read_xlsx_table_range(
        file_path,
        "Лист1",
        "C3:D10",
        entity_type="group",
    )
    students = service.read_xlsx_table_range(
        file_path,
        "Лист1",
        "L12:O16",
        entity_type="student",
    )
    second_groups = service.read_xlsx_table_range(
        file_path,
        "Лист2",
        "C5:D11",
        entity_type="group",
    )

    assert first_groups.headers == ("Группы", "Специальности")
    assert first_groups.rows == [
        {"Группы": "СМ1-21Б", "Специальности": 123},
        {"Группы": "СМ2-51Б", "Специальности": None},
        {"Группы": "СМ3-21", "Специальности": 288},
        {"Группы": "СМ4-41", "Специальности": 335},
        {"Группы": "СМ5-21Б", "Специальности": None},
        {"Группы": "СМ6-21Б", "Специальности": 433},
        {"Группы": "СМ7-31Б", "Специальности": 488},
    ]
    assert first_groups.errors == ["Missing required headers: name."]
    assert first_groups.is_valid is False

    assert students.headers == ("Фамилия", "Имя", "Группа", "Почта")
    assert students.rows == [
        {
            "Фамилия": "Иванов",
            "Имя": "Иван",
            "Группа": "СМ1-21Б",
            "Почта": "a@test.ru",
        },
        {
            "Фамилия": "Петров",
            "Имя": "Петр",
            "Группа": "СМ6-21Б",
            "Почта": None,
        },
        {
            "Фамилия": "Сидоров",
            "Имя": "Сидор",
            "Группа": "СМ4-41",
            "Почта": "v@test.ru",
        },
        {
            "Фамилия": "Горшенев",
            "Имя": "Миша",
            "Группа": "СМ7-31Б",
            "Почта": "e@test.ru",
        },
    ]
    assert students.errors == [
        "Missing required headers: group_id, surname, first_name."
    ]
    assert students.is_valid is False

    assert second_groups.headers == ("Группы", "Специальности")
    assert second_groups.rows == [
        {"Группы": "ИУ1-21Б", "Специальности": 123},
        {"Группы": "ИУ2-51Б", "Специальности": None},
        {"Группы": "ИУ3-21", "Специальности": 184},
        {"Группы": "ИУ5-21Б", "Специальности": 297},
        {"Группы": "ИУ6-21Б", "Специальности": 340},
    ]
    assert second_groups.errors == ["Missing required headers: name."]
    assert second_groups.is_valid is False


def test_data_normalizer_maps_student_row_and_separates_group_reference():
    """
    Нормализатор раскладывает строку студента на прямые поля и ссылку на группу.
    """
    normalizer = DataNormalizer()
    extracted_table = ExtractedTable(
        sheet="Лист1",
        range="L12:O16",
        entity_type="students",
        headers=("Фамилия", "Имя", "Группа", "Почта", "Лишнее"),
        rows=[
            {
                "Фамилия": "Иванов",
                "Имя": "Иван",
                "Группа": "СМ1-21Б",
                "Почта": "a@test.ru",
                "Лишнее": "заметка",
            }
        ],
    )

    normalized_rows = normalizer.normalize_table(extracted_table)

    assert len(normalized_rows) == 1
    assert normalized_rows[0].data == {
        "surname": "Иванов",
        "first_name": "Иван",
        "bmstu_email": "a@test.ru",
    }
    assert normalized_rows[0].references == {"group": {"name": "СМ1-21Б"}}
    assert normalized_rows[0].unmapped == {"Лишнее": "заметка"}
    assert normalized_rows[0].errors == []


def test_data_normalizer_warns_about_conflicting_alias_values():
    """
    Если две колонки маппятся в одно поле с разными значениями, нормализатор оставляет warning.
    """
    normalizer = DataNormalizer()

    normalized_row = normalizer.normalize_row(
        "students",
        {
            "surname": "Иванов",
            "Фамилия": "Петров",
            "Имя": "Иван",
        },
        source_sheet="Лист1",
        source_range="A1:C2",
        source_row_number=2,
    )

    assert normalized_row.data["surname"] == "Иванов"
    assert normalized_row.data["first_name"] == "Иван"
    assert normalized_row.warnings == [
        "Conflicting values for field surname: 'Иванов' and 'Петров'."
    ]
    assert normalized_row.errors == []


def test_xlsx_config_separates_smart_and_strict_entity_types():
    """
    XLSX-конфиг явно разделяет сущности умного и строгого импорта.
    """
    assert SMART_IMPORT_ENTITY_TYPES == ("groups", "schedules", "students")
    assert STRICT_IMPORT_ENTITY_TYPES == (
        "schedule_group_links",
        "lessons",
        "attendances",
        "marks",
        "comments",
    )


def test_data_normalizer_rejects_strict_entity_type():
    """
    Нормализатор не должен пытаться обрабатывать сущности строгого импорта.
    """
    normalizer = DataNormalizer()
    extracted_table = ExtractedTable(
        sheet="Лист1",
        range="A1:F2",
        entity_type="attendances",
        headers=("student_id", "lesson_id", "is_visited"),
        rows=[{"student_id": 1, "lesson_id": 2, "is_visited": True}],
    )

    with pytest.raises(
        ValueError,
        match="is not supported for smart import",
    ):
        normalizer.normalize_table(extracted_table)


def test_data_resolver_resolves_group_reference_to_group_id():
    """
    Резолвер может подставить `group_id`, если нормализатор сохранил ссылку по имени группы.
    """
    normalizer = DataNormalizer()
    extracted_table = ExtractedTable(
        sheet="Лист1",
        range="L12:O16",
        entity_type="students",
        headers=("Фамилия", "Имя", "Группа"),
        rows=[{"Фамилия": "Иванов", "Имя": "Иван", "Группа": "СМ1-21Б"}],
    )

    normalized_row = normalizer.normalize_table(extracted_table)[0]
    resolver = DataResolver(
        reference_resolvers={
            "group": lambda criteria: {"group_id": 15}
            if criteria == {"name": "СМ1-21Б"}
            else None
        }
    )

    resolved_row = resolver.resolve_rows([normalized_row])[0]

    assert resolved_row.data == {
        "surname": "Иванов",
        "first_name": "Иван",
        "group_id": 15,
    }
    assert resolved_row.resolved_references == {"group": {"group_id": 15}}
    assert resolved_row.unresolved_references == {}
    assert resolved_row.errors == []


def test_data_resolver_keeps_unresolved_reference_when_callback_is_missing():
    """
    Если callback для ссылки не настроен, резолвер не падает и помечает ссылку как неразрешенную.
    """
    resolver = DataResolver()
    row = DataNormalizer().normalize_row(
        "students",
        {"Фамилия": "Иванов", "Имя": "Иван", "Группа": "СМ1-21Б"},
        source_sheet="Лист1",
        source_range="A1:C2",
        source_row_number=2,
    )

    resolved_row = resolver.resolve_rows([row])[0]

    assert resolved_row.data == {"surname": "Иванов", "first_name": "Иван"}
    assert resolved_row.unresolved_references == {"group": {"name": "СМ1-21Б"}}
    assert resolved_row.warnings == [
        "No resolver was configured for reference 'group'."
    ]
    assert resolved_row.errors == []


def test_data_resolver_reports_unresolved_reference_when_callback_returns_none():
    """
    Если callback не смог разрешить ссылку, это считается ошибкой строки.
    """
    row = DataNormalizer().normalize_row(
        "students",
        {"Фамилия": "Иванов", "Имя": "Иван", "Группа": "СМ1-21Б"},
        source_sheet="Лист1",
        source_range="A1:C2",
        source_row_number=2,
    )
    resolver = DataResolver(reference_resolvers={"group": lambda criteria: None})

    resolved_row = resolver.resolve_rows([row])[0]

    assert resolved_row.unresolved_references == {"group": {"name": "СМ1-21Б"}}
    assert resolved_row.errors == [
        "Could not resolve reference 'group' with criteria {'name': 'СМ1-21Б'}."
    ]
    assert resolved_row.is_valid is False


def test_import_processor_builds_group_create_payloads():
    """
    Процессор собирает итоговые payload-словари, готовые для create-схем.
    """
    processor = ImportProcessor()
    extracted_table = ExtractedTable(
        sheet="Лист1",
        range="C3:D5",
        entity_type="groups",
        headers=("Группы", "Специальности"),
        rows=[
            {"Группы": "СМ1-21Б", "Специальности": "24.03.01_Информатика"},
            {"Группы": "СМ2-51Б", "Специальности": None},
        ],
    )

    result = processor.process_table(extracted_table)

    assert result.entity_type == "groups"
    assert result.create_payloads == [
        {"name": "СМ1-21Б", "speciality": "24.03.01_Информатика"},
        {"name": "СМ2-51Б"},
    ]
    assert result.errors == []
    assert result.is_valid is True


def test_import_processor_reports_create_schema_validation_error():
    """
    Если после нормализации не хватает обязательных полей create-схемы, процессор возвращает понятную ошибку.
    """
    processor = ImportProcessor()
    extracted_table = ExtractedTable(
        sheet="Лист1",
        range="D5:D6",
        entity_type="groups",
        headers=("Специальности",),
        rows=[{"Специальности": "24.03.01_Информатика"}],
    )

    result = processor.process_table(extracted_table)

    assert result.create_payloads == []
    assert result.errors == ["row 2: name: Field required"]
    assert result.is_valid is False


def test_data_processor_builds_header_bindings_and_row_trace():
    """
    Обработчик данных сохраняет карту заголовков и построчную трассировку распознавания.
    """
    processor = DataProcessor(
        data_resolver=DataResolver(
            reference_resolvers={"group": lambda criteria: {"group_id": 15}}
        )
    )
    extracted_table = ExtractedTable(
        sheet="Лист1",
        range="L12:O16",
        entity_type="students",
        headers=("Фамилия", "Имя", "Группа", "Почта", "Лишнее"),
        rows=[
            {
                "Фамилия": "Иванов",
                "Имя": "Иван",
                "Группа": "СМ1-21Б",
                "Почта": "a@test.ru",
                "Лишнее": "заметка",
            }
        ],
    )

    result = processor.process_table(extracted_table)

    assert result.entity_type == "students"
    assert result.header_bindings == [
        HeaderBinding("Фамилия", "фамилия", "direct", "surname"),
        HeaderBinding("Имя", "имя", "direct", "first_name"),
        HeaderBinding("Группа", "группа", "reference", "group.name"),
        HeaderBinding("Почта", "почта", "direct", "bmstu_email"),
        HeaderBinding("Лишнее", "лишнее", "unmapped", None),
    ]
    assert result.create_payloads == [
        {
            "surname": "Иванов",
            "first_name": "Иван",
            "group_id": 15,
            "bmstu_email": "a@test.ru",
        }
    ]
    assert result.rows == [
        ProcessedRow(
            source_row_number=2,
            source_values={
                "Фамилия": "Иванов",
                "Имя": "Иван",
                "Группа": "СМ1-21Б",
                "Почта": "a@test.ru",
                "Лишнее": "заметка",
            },
            normalized_data={
                "surname": "Иванов",
                "first_name": "Иван",
                "bmstu_email": "a@test.ru",
            },
            references={"group": {"name": "СМ1-21Б"}},
            resolved_data={
                "surname": "Иванов",
                "first_name": "Иван",
                "bmstu_email": "a@test.ru",
                "group_id": 15,
            },
            resolved_references={"group": {"group_id": 15}},
            unresolved_references={},
            unmapped={"Лишнее": "заметка"},
            warnings=[],
            errors=[],
            create_payload={
                "surname": "Иванов",
                "first_name": "Иван",
                "group_id": 15,
                "bmstu_email": "a@test.ru",
            },
        )
    ]
    assert result.errors == []
    assert result.is_valid is True


def test_data_processor_preserves_extracted_table_errors():
    """
    Ошибки диапазона верхнего уровня не теряются после обработки данных.
    """
    processor = DataProcessor()
    extracted_table = ExtractedTable(
        sheet="Лист1",
        range="D5:E6",
        entity_type="groups",
        headers=("Специальности", "Комментарий"),
        rows=[{"Специальности": "24.03.01_Информатика", "Комментарий": "заметка"}],
        errors=["Missing required headers: name."],
        warnings=["Header Комментарий is not recognized."],
    )

    result = processor.process_table(extracted_table)

    assert result.warnings == ["Header Комментарий is not recognized."]
    assert result.errors == ["Missing required headers: name.", "row 2: name: Field required"]
    assert result.create_payloads == []
    assert result.is_valid is False


def test_import_export_service_reads_strict_table_with_standard_headers(tmp_path):
    """
    Строгий импорт принимает таблицу из произвольного диапазона, если ее формат стандартный.
    """
    service = ImportExportService()
    file_path = tmp_path / "strict_attendance.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    headers = XLSX_COLUMNS_BY_SHEET["attendances"]

    for index, header in enumerate(headers, start=3):
        worksheet.cell(row=7, column=index, value=header)

    worksheet.cell(row=8, column=3, value=None)
    worksheet.cell(row=8, column=4, value=11)
    worksheet.cell(row=8, column=5, value=22)
    worksheet.cell(row=8, column=6, value=True)
    worksheet.cell(row=8, column=7, value=None)
    worksheet.cell(row=8, column=8, value=None)
    workbook.save(file_path)

    result = service.read_strict_xlsx_table(
        file_path,
        "Sheet1",
        "C7:H8",
        entity_type="attendance",
    )

    assert result.entity_type == "attendances"
    assert result.create_payloads == [
        {"student_id": 11, "lesson_id": 22, "is_visited": True}
    ]
    assert result.errors == []
    assert result.is_valid is True


def test_import_export_service_processes_smart_table_with_recognition_map(tmp_path):
    """
    Верхний сервис возвращает не только payload-ы, но и карту распознавания smart-таблицы.
    """
    service = ImportExportService()
    file_path = tmp_path / "smart_processing.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["C4"] = "Фамилия"
    worksheet["D4"] = "Имя"
    worksheet["E4"] = "Группа"
    worksheet["F4"] = "Почта"
    worksheet["C5"] = "Иванов"
    worksheet["D5"] = "Иван"
    worksheet["E5"] = "СМ1-21Б"
    worksheet["F5"] = "a@test.ru"
    workbook.save(file_path)

    result = service.process_smart_xlsx_table(
        file_path,
        "Sheet1",
        "C4:F5",
        entity_type="student",
    )

    assert isinstance(result, DataProcessingResult)
    assert result.entity_type == "students"
    assert [binding.target_path for binding in result.header_bindings] == [
        "surname",
        "first_name",
        "group.name",
        "bmstu_email",
    ]
    assert result.rows[0].normalized_data == {
        "surname": "Иванов",
        "first_name": "Иван",
        "bmstu_email": "a@test.ru",
    }
    assert result.rows[0].references == {"group": {"name": "СМ1-21Б"}}
    assert result.create_payloads == []
    assert result.warnings == [
        "No resolver was configured for reference 'group'."
    ]
    assert result.errors == [
        "Missing required headers: group_id, surname, first_name.",
        "row 2: group_id: Field required",
    ]


def test_strict_import_processor_warns_when_header_order_differs(tmp_path):
    """
    Строгий импорт допускает перестановку правильных колонок, но оставляет warning.
    """
    processor = StrictImportProcessor()
    extracted_table = ExtractedTable(
        sheet="Sheet1",
        range="B2:G3",
        entity_type="attendances",
        headers=("lesson_id", "student_id", "is_visited", "id", "created_at", "updated_at"),
        rows=[
            {
                "lesson_id": 7,
                "student_id": 5,
                "is_visited": True,
                "id": None,
                "created_at": None,
                "updated_at": None,
            }
        ],
    )

    result = processor.process_table(extracted_table)

    assert result.create_payloads == [
        {"student_id": 5, "lesson_id": 7, "is_visited": True}
    ]
    assert result.warnings == ["Header order differs from the standard XLSX format."]
    assert result.errors == []
    assert result.is_valid is True


def test_import_export_service_rejects_strict_table_with_unknown_headers(tmp_path):
    """
    Строгий импорт отклоняет таблицу, если в ней есть лишние колонки.
    """
    service = ImportExportService()
    file_path = tmp_path / "strict_bad_comment.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    headers = ("student_id", "lesson_id", "data", "Комментарий")

    for index, header in enumerate(headers, start=2):
        worksheet.cell(row=5, column=index, value=header)

    worksheet.cell(row=6, column=2, value=1)
    worksheet.cell(row=6, column=3, value=2)
    worksheet.cell(row=6, column=4, value="text")
    worksheet.cell(row=6, column=5, value="extra")
    workbook.save(file_path)

    result = service.read_strict_xlsx_table(
        file_path,
        "Sheet1",
        "B5:E6",
        entity_type="comment",
    )

    assert result.entity_type == "comments"
    assert result.create_payloads == []
    assert result.errors == [
        "Missing strict headers: id, created_at, updated_at.",
        "Unknown strict headers: Комментарий.",
    ]
    assert result.is_valid is False


def test_io_tools_emit_basic_logs(tmp_path, caplog):
    """
    Слой io_tools пишет базовые служебные логи при экспорте и импорте.
    """
    service = ImportExportService()
    file_path = tmp_path / "logged_io.xlsx"
    payload = {
        "groups": [{"id": 1, "name": "IU7-11"}],
    }

    with caplog.at_level(logging.DEBUG):
        service.export_to_xlsx(payload, file_path)
        service.import_from_xlsx(file_path)

    assert "XLSX export requested" in caplog.text
    assert "XlsxExporter export started" in caplog.text
    assert "XlsxExporter export finished" in caplog.text
    assert "XLSX import requested" in caplog.text
    assert "XlsxImporter import started" in caplog.text
    assert "XlsxImporter import finished" in caplog.text


def test_xlsx_round_trip_exports_all_database_entities(db_session):
    """
    Проверяет полный круг экспорта данных из БД в XLSX и чтения файла обратно.

    Тест создает по три записи каждой сущности, сохраняет экспорт в папку `logs`,
    затем считывает этот файл через импортёр и сравнивает содержимое всех листов
    с теми данными, которые реально были записаны в базу до экспорта.

    :param db_session: SQLAlchemy-сессия тестовой базы данных.
    :return: `None`. Тест падает, если хотя бы один лист экспортирован или прочитан
        обратно с потерей данных либо нарушением структуры.
    """
    base_time = datetime(2026, 5, 10, 20, 0, 0)

    groups = [
        Group(
            name=f"IU7-9{index}",
            speciality=f"09.03.0{index + 1}_Informatics",
            created_at=base_time.replace(minute=index),
            updated_at=base_time.replace(minute=index, second=10),
        )
        for index in range(1, 4)
    ]
    db_session.add_all(groups)
    db_session.flush()

    schedules = [
        Schedule(
            odd_or_even="odd" if index % 2 else "even",
            type=f"type_{index}",
            is_assessment=index % 2 == 0,
            day=f"day_{index}",
            time=time(8 + index, 15),
            created_at=base_time.replace(minute=10 + index),
            updated_at=base_time.replace(minute=10 + index, second=10),
        )
        for index in range(1, 4)
    ]
    db_session.add_all(schedules)
    db_session.flush()

    students = [
        Student(
            group_id=groups[index - 1].id,
            surname=f"Surname{index}",
            first_name=f"Name{index}",
            patronymic=f"Patronymic{index}",
            personal_data=f"PD{index:03d}",
            bmstu_email=f"student{index}@bmstu.ru",
            created_at=base_time.replace(minute=20 + index),
            updated_at=base_time.replace(minute=20 + index, second=10),
        )
        for index in range(1, 4)
    ]
    db_session.add_all(students)
    db_session.flush()

    schedule_group_links = [
        ScheduleGroupLink(
            group_id=groups[index - 1].id,
            schedule_id=schedules[index - 1].id,
            created_at=base_time.replace(minute=30 + index),
            updated_at=base_time.replace(minute=30 + index, second=10),
        )
        for index in range(1, 4)
    ]
    db_session.add_all(schedule_group_links)
    db_session.flush()

    lessons = [
        Lesson(
            schedule_id=schedules[index - 1].id,
            topic=f"Topic {index}",
            date=date(2026, 5, 10 + index),
            created_at=base_time.replace(minute=40 + index),
            updated_at=base_time.replace(minute=40 + index, second=10),
        )
        for index in range(1, 4)
    ]
    db_session.add_all(lessons)
    db_session.flush()

    attendances = [
        Attendance(
            student_id=students[index - 1].id,
            lesson_id=lessons[index - 1].id,
            is_visited=index % 2 == 1,
            created_at=base_time.replace(minute=50 + index),
            updated_at=base_time.replace(minute=50 + index, second=10),
        )
        for index in range(1, 4)
    ]
    db_session.add_all(attendances)
    db_session.flush()

    marks = [
        Mark(
            student_id=students[index - 1].id,
            lesson_id=lessons[index - 1].id,
            data=2 + index,
            created_at=base_time.replace(hour=21, minute=index),
            updated_at=base_time.replace(hour=21, minute=index, second=10),
        )
        for index in range(1, 4)
    ]
    db_session.add_all(marks)
    db_session.flush()

    comments = [
        Comment(
            student_id=students[index - 1].id,
            lesson_id=lessons[index - 1].id,
            data=f"Comment {index}",
            created_at=base_time.replace(hour=21, minute=10 + index),
            updated_at=base_time.replace(hour=21, minute=10 + index, second=10),
        )
        for index in range(1, 4)
    ]
    db_session.add_all(comments)
    db_session.flush()

    expected_rows = {
        "groups": [_normalize_db_row(row, "groups") for row in groups],
        "schedules": [_normalize_db_row(row, "schedules") for row in schedules],
        "students": [_normalize_db_row(row, "students") for row in students],
        "schedule_group_links": [
            _normalize_db_row(row, "schedule_group_links")
            for row in schedule_group_links
        ],
        "lessons": [_normalize_db_row(row, "lessons") for row in lessons],
        "attendances": [_normalize_db_row(row, "attendances") for row in attendances],
        "marks": [_normalize_db_row(row, "marks") for row in marks],
        "comments": [_normalize_db_row(row, "comments") for row in comments],
    }

    logs_dir = _logs_dir()
    logs_dir.mkdir(parents=True, exist_ok=True)
    file_path = logs_dir / "test_io_tools_round_trip.xlsx"

    orm_service = OrmService(db_session, auto_commit=False)
    importer = ImportExportService()
    exported_path = orm_service.export_to_xlsx(
        [sheet_name[:-1] if sheet_name.endswith("s") else sheet_name for sheet_name in XLSX_SHEETS_ORDER],
        file_path,
    )
    imported_rows = importer.import_from_xlsx(exported_path)

    assert exported_path == file_path
    assert tuple(imported_rows) == XLSX_SHEETS_ORDER

    for sheet_name in XLSX_SHEETS_ORDER:
        assert _normalize_imported_rows(
            imported_rows[sheet_name],
            expected_rows[sheet_name],
        ) == _sorted_rows(
            expected_rows[sheet_name]
        )
