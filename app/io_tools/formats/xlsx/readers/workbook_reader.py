"""
Чтение стандартного XLSX-формата приложения с диагностикой.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.io_tools.engine.operation_result import TabularImportResult
from app.io_tools.xlsx_config import (
    XLSX_COLUMNS_BY_SHEET,
    XLSX_REQUIRED_COLUMNS_BY_SHEET,
    XLSX_SHEETS_ORDER,
)


logger = logging.getLogger(__name__)


class StandardWorkbookReader:
    """
    Читает XLSX-книгу стандартного внутреннего формата приложения.

    В отличие от старого сценария, reader не обязан сразу выбрасывать
    исключение. Он сначала собирает данные и диагностические сообщения,
    чтобы верхний слой сам решил, продолжать ли работу.
    """

    def read(self, file_path: str | Path) -> TabularImportResult:
        """
        Считывает книгу стандартного XLSX-формата.

        :param file_path: Путь к XLSX-файлу.
        :return: Результат чтения с данными и диагностикой.
        """
        logger.info(
            "StandardWorkbookReader read started. source=%s.",
            file_path,
        )
        workbook = load_workbook(file_path)
        result = TabularImportResult()
        sheet_names = workbook.sheetnames

        self._validate_sheet_names(sheet_names, result)
        if not result.is_valid:
            logger.warning(
                "StandardWorkbookReader found workbook-level errors. source=%s errors=%s.",
                file_path,
                result.errors,
            )
            return result

        for sheet_name in self._ordered_sheet_names(sheet_names):
            worksheet = workbook[sheet_name]
            result.data[sheet_name] = self._read_sheet(
                sheet_name,
                worksheet,
                result,
            )

        logger.info(
            "StandardWorkbookReader read finished. source=%s sheets_count=%s errors=%s.",
            file_path,
            len(result.data),
            len(result.errors),
        )
        return result

    def _validate_sheet_names(
        self,
        sheet_names: list[str],
        result: TabularImportResult,
    ) -> None:
        """
        Проверяет, что книга не содержит неизвестных листов.

        :param sheet_names: Имена листов книги.
        :param result: Накопитель результата чтения.
        :return: `None`.
        """
        unknown_sheet_names = [
            sheet_name
            for sheet_name in sheet_names
            if sheet_name not in XLSX_COLUMNS_BY_SHEET
        ]
        if unknown_sheet_names:
            joined = ", ".join(unknown_sheet_names)
            message = f"Неизвестные листы в XLSX-файле: {joined}."
            logger.error("StandardWorkbookReader found unknown sheets: %s.", joined)
            result.add_error(message)

    @staticmethod
    def _ordered_sheet_names(sheet_names: list[str]) -> list[str]:
        """
        Возвращает листы в согласованном порядке.

        :param sheet_names: Фактические имена листов книги.
        :return: Упорядоченный список листов.
        """
        ordered_sheet_names = [
            sheet_name
            for sheet_name in XLSX_SHEETS_ORDER
            if sheet_name in sheet_names
        ]
        ordered_sheet_names.extend(
            sheet_name
            for sheet_name in sheet_names
            if sheet_name not in ordered_sheet_names
        )
        return ordered_sheet_names

    def _read_sheet(
        self,
        sheet_name: str,
        worksheet: Any,
        result: TabularImportResult,
    ) -> list[dict[str, object]]:
        """
        Читает один лист стандартной XLSX-книги.

        :param sheet_name: Имя листа.
        :param worksheet: Лист openpyxl.
        :param result: Накопитель результата чтения.
        :return: Список строк листа в виде словарей.
        """
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            self._validate_headers(sheet_name, (), result)
            logger.debug(
                "StandardWorkbookReader read empty sheet=%s.",
                sheet_name,
            )
            return []

        headers = tuple(
            str(header).strip()
            for header in rows[0]
            if header is not None and str(header).strip()
        )
        error_count_before = len(result.errors)
        self._validate_headers(sheet_name, headers, result)
        if len(result.errors) > error_count_before:
            return []

        imported_rows: list[dict[str, object]] = []
        for raw_row in rows[1:]:
            row_data = {
                header: raw_row[index] if index < len(raw_row) else None
                for index, header in enumerate(headers)
            }
            if self._is_empty_row(row_data):
                continue
            imported_rows.append(row_data)

        logger.debug(
            "StandardWorkbookReader read sheet=%s rows_count=%s headers=%s.",
            sheet_name,
            len(imported_rows),
            headers,
        )
        return imported_rows

    @staticmethod
    def _is_empty_row(row_data: dict[str, object]) -> bool:
        """
        Проверяет, содержит ли строка хотя бы одно полезное значение.

        :param row_data: Словарь значений строки.
        :return: `True`, если все значения пустые.
        """
        return all(value is None for value in row_data.values())

    def _validate_headers(
        self,
        sheet_name: str,
        headers: tuple[str, ...],
        result: TabularImportResult,
    ) -> None:
        """
        Проверяет обязательные заголовки стандартного листа.

        :param sheet_name: Имя листа.
        :param headers: Заголовки, найденные в файле.
        :param result: Накопитель результата чтения.
        :return: `None`.
        """
        required_headers = XLSX_REQUIRED_COLUMNS_BY_SHEET.get(sheet_name, ())
        missing_headers = [
            header
            for header in required_headers
            if header not in headers
        ]
        if missing_headers:
            joined = ", ".join(missing_headers)
            message = (
                f"В листе {sheet_name} отсутствуют обязательные колонки: {joined}."
            )
            logger.error(
                "StandardWorkbookReader missing required headers. sheet=%s headers=%s.",
                sheet_name,
                joined,
            )
            result.add_error(message, sheet_name=sheet_name)
