"""
Импорт структурированных данных приложения из формата XLSX.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.io_tools.xlsx_config import (
    XLSX_COLUMNS_BY_SHEET,
    XLSX_REQUIRED_COLUMNS_BY_SHEET,
    XLSX_SHEETS_ORDER,
)


logger = logging.getLogger(__name__)


class XlsxImporter:
    """
    Каркас импортера данных из XLSX.

    Класс добавлен как часть общей архитектуры `io_tools`. Реализация импорта
    будет развиваться отдельно, когда будет согласован формат входного файла
    и правила загрузки данных в базу.
    """

    def import_data(self, file_path: str | Path) -> dict[str, list[dict[str, object]]]:
        """
        Читает XLSX-файл и возвращает структурированные данные.

        :param file_path: Путь к XLSX-файлу импорта.
        :return: Структурированные данные по листам.
        :raises ValueError: Если файл содержит неизвестные листы или в листе не
            хватает обязательных колонок.
        """
        logger.info("XlsxImporter import started. source=%s.", file_path)
        workbook = load_workbook(file_path)
        workbook_sheet_names = workbook.sheetnames
        logger.debug(
            "XlsxImporter workbook opened. sheets=%s.",
            workbook_sheet_names,
        )
        self._validate_sheet_names(workbook_sheet_names)

        imported_data: dict[str, list[dict[str, object]]] = {}
        ordered_sheet_names = [
            sheet_name
            for sheet_name in XLSX_SHEETS_ORDER
            if sheet_name in workbook_sheet_names
        ]
        ordered_sheet_names.extend(
            sheet_name
            for sheet_name in workbook_sheet_names
            if sheet_name not in ordered_sheet_names
        )

        for sheet_name in ordered_sheet_names:
            worksheet = workbook[sheet_name]
            imported_data[sheet_name] = self._read_sheet(sheet_name, worksheet)

        logger.info(
            "XlsxImporter import finished. source=%s sheets_count=%s.",
            file_path,
            len(imported_data),
        )
        return imported_data

    @staticmethod
    def _validate_sheet_names(sheet_names: list[str]) -> None:
        """
        Проверяет, что файл не содержит неизвестных листов.

        :param sheet_names: Имена листов книги.
        :return: `None`.
        :raises ValueError: Если найден хотя бы один неизвестный лист.
        """
        unknown_sheet_names = [
            sheet_name
            for sheet_name in sheet_names
            if sheet_name not in XLSX_COLUMNS_BY_SHEET
        ]
        if unknown_sheet_names:
            joined = ", ".join(unknown_sheet_names)
            logger.error("XlsxImporter found unknown sheets: %s.", joined)
            raise ValueError(f"Неизвестные листы в XLSX-файле: {joined}.")

    def _read_sheet(
        self,
        sheet_name: str,
        worksheet: Any,
    ) -> list[dict[str, object]]:
        """
        Читает один лист книги и возвращает его строки как словари.

        :param sheet_name: Имя листа.
        :param worksheet: Лист openpyxl.
        :return: Список строк листа в виде словарей.
        :raises ValueError: Если в листе не хватает обязательных колонок.
        """
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            self._validate_headers(sheet_name, ())
            logger.debug("XlsxImporter read empty sheet=%s.", sheet_name)
            return []

        raw_headers = rows[0]
        headers = tuple(
            str(header).strip()
            for header in raw_headers
            if header is not None and str(header).strip()
        )
        self._validate_headers(sheet_name, headers)

        imported_rows: list[dict[str, object]] = []
        for raw_row in rows[1:]:
            row_data = {
                header: raw_row[index] if index < len(raw_row) else None
                for index, header in enumerate(headers)
            }
            if self._is_empty_row(row_data):
                continue
            imported_rows.append(row_data)

        logger.debug(
            "XlsxImporter read sheet=%s rows_count=%s headers=%s.",
            sheet_name,
            len(imported_rows),
            headers,
        )
        return imported_rows

    @staticmethod
    def _is_empty_row(row_data: dict[str, object]) -> bool:
        """
        Проверяет, является ли строка Excel пустой.

        :param row_data: Словарь значений строки.
        :return: `True`, если строка фактически пустая, иначе `False`.
        """
        return all(value is None for value in row_data.values())

    @staticmethod
    def _validate_headers(sheet_name: str, headers: tuple[str, ...]) -> None:
        """
        Проверяет обязательные колонки листа.

        :param sheet_name: Имя листа.
        :param headers: Колонки, считанные из файла.
        :return: `None`.
        :raises ValueError: Если в листе не хватает обязательных колонок.
        """
        required_headers = XLSX_REQUIRED_COLUMNS_BY_SHEET.get(sheet_name, ())
        missing_headers = [
            header
            for header in required_headers
            if header not in headers
        ]
        if missing_headers:
            joined = ", ".join(missing_headers)
            logger.error(
                "XlsxImporter missing required headers. sheet=%s headers=%s.",
                sheet_name,
                joined,
            )
            raise ValueError(
                f"В листе {sheet_name} отсутствуют обязательные колонки: "
                f"{joined}."
            )
