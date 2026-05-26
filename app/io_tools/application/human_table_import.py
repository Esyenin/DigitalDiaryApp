"""
Human table import для "человеческих" Excel-файлов.

Поддерживает файлы, где:
- таблица групп может называться по-русски: группа | специальность
- таблица студентов может лежать рядом: фамилия | имя | ...
- таблица оценок может быть отдельной: студент | оценки

Если в файле есть оценки, но нет урока, importer создает техническое
расписание и техническое занятие.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

from config import settings
from app.models import (
    Group,
    Lesson,
    Mark,
    Schedule,
    ScheduleGroupLink,
    Student,
)


GROUP_HEADERS = {
    "группа": "group_name",
    "group": "group_name",
    "название группы": "group_name",
    "специальность": "speciality",
    "speciality": "speciality",
}

STUDENT_HEADERS = {
    "фамилия": "surname",
    "surname": "surname",
    "last name": "surname",
    "имя": "first_name",
    "first name": "first_name",
    "name": "first_name",
    "группа": "group_name",
    "group": "group_name",
    "почта": "bmstu_email",
    "email": "bmstu_email",
}

MARK_HEADERS = {
    "студент": "student",
    "student": "student",
    "фамилия": "student",
    "оценка": "mark",
    "оценки": "mark",
    "mark": "mark",
    "marks": "mark",
    "grade": "mark",
}


@dataclass(slots=True)
class HumanTableImportResult:
    created: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not self.errors


@dataclass(slots=True)
class FoundTable:
    sheet_name: str
    header_row: int
    header_map: dict[int, str]
    table_type: str


class HumanTableImporter:
    def __init__(self, session: Session) -> None:
        self.session = session
        self.group_name_to_id: dict[str, int] = {}
        self.student_name_to_id: dict[str, int] = {}
        self.student_surname_to_ids: dict[str, list[int]] = {}

    def import_workbook(self, file_path: str | Path) -> HumanTableImportResult:
        workbook = load_workbook(Path(file_path), data_only=True)

        result = HumanTableImportResult()

        tables = self._find_tables(workbook)

        if not tables:
            result.errors.append(
                "Не найдено ни одной поддерживаемой human-table таблицы.\n\n"
                "Ожидаются заголовки вроде: группа, специальность, фамилия, имя, студент, оценки."
            )
            return result

        try:
            # 1. Сначала группы
            for table in tables:
                if table.table_type == "groups":
                    self._import_groups(workbook, table, result)

            # 2. Потом студенты
            for table in tables:
                if table.table_type == "students":
                    self._import_students(workbook, table, result)

            # 3. Потом оценки
            for table in tables:
                if table.table_type == "marks":
                    self._import_marks(workbook, table, result)

            if result.errors:
                self.session.rollback()
                return result

            self.session.commit()
            return result

        except Exception as exc:
            self.session.rollback()
            result.errors.append(f"Unexpected human table import error: {exc}")
            return result

    def _find_tables(self, workbook: Any) -> list[FoundTable]:
        tables: list[FoundTable] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row_index in range(1, sheet.max_row + 1):
                row_values = [
                    sheet.cell(row=row_index, column=col_index).value
                    for col_index in range(1, sheet.max_column + 1)
                ]

                group_header_map = self._match_headers(row_values, GROUP_HEADERS)
                student_header_map = self._match_headers(row_values, STUDENT_HEADERS)
                mark_header_map = self._match_headers(row_values, MARK_HEADERS)

                if {"group_name", "speciality"} <= set(group_header_map.values()):
                    tables.append(
                        FoundTable(
                            sheet_name=sheet_name,
                            header_row=row_index,
                            header_map=group_header_map,
                            table_type="groups",
                        )
                    )

                if {"surname", "first_name"} <= set(student_header_map.values()):
                    tables.append(
                        FoundTable(
                            sheet_name=sheet_name,
                            header_row=row_index,
                            header_map=student_header_map,
                            table_type="students",
                        )
                    )

                if {"student", "mark"} <= set(mark_header_map.values()):
                    tables.append(
                        FoundTable(
                            sheet_name=sheet_name,
                            header_row=row_index,
                            header_map=mark_header_map,
                            table_type="marks",
                        )
                    )

        return tables

    @staticmethod
    def _match_headers(
        row_values: list[Any],
        aliases: dict[str, str],
    ) -> dict[int, str]:
        header_map: dict[int, str] = {}

        for zero_based_index, value in enumerate(row_values):
            if value is None:
                continue

            normalized = str(value).strip().lower()
            canonical = aliases.get(normalized)

            if canonical is not None:
                header_map[zero_based_index + 1] = canonical

        return header_map

    def _read_table_rows(
            self,
            workbook: Any,
            table: FoundTable,
    ) -> list[tuple[int, dict[str, Any]]]:
        sheet = workbook[table.sheet_name]
        rows: list[tuple[int, dict[str, Any]]] = []

        empty_rows_in_a_row = 0

        for row_index in range(table.header_row + 1, sheet.max_row + 1):
            row_data: dict[str, Any] = {}

            for col_index, field_name in table.header_map.items():
                value = sheet.cell(row=row_index, column=col_index).value

                if value is None or value == "":
                    continue

                row_data[field_name] = value

            if not row_data:
                empty_rows_in_a_row += 1

                if empty_rows_in_a_row >= 2:
                    break

                continue

            empty_rows_in_a_row = 0

            if not self._row_matches_table_type(table.table_type, row_data):
                break

            rows.append((row_index, row_data))

        return rows
    @staticmethod
    def _row_matches_table_type(
        table_type: str,
        row_data: dict[str, Any],
    ) -> bool:
        fields = set(row_data)

        if table_type == "groups":
            return "group_name" in fields

        if table_type == "students":
            return bool({"surname", "first_name"} & fields)

        if table_type == "marks":
            return bool({"student", "mark"} & fields)

        return True
    def _import_groups(
        self,
        workbook: Any,
        table: FoundTable,
        result: HumanTableImportResult,
    ) -> None:
        created_count = result.created.get("groups", 0)

        for row_index, row_data in self._read_table_rows(workbook, table):
            group_name = self._clean_string(row_data.get("group_name"))
            speciality = self._clean_string(row_data.get("speciality"))

            if not group_name:
                result.warnings.append(
                    f"{table.sheet_name}, row {row_index}: group name is empty, skipped."
                )
                continue

            group = self._get_or_create_group(group_name, speciality)
            self.group_name_to_id[group_name] = group.id

            if group.created_at == group.updated_at:
                created_count += 1

        result.created["groups"] = created_count

    def _import_students(
        self,
        workbook: Any,
        table: FoundTable,
        result: HumanTableImportResult,
    ) -> None:
        created_count = result.created.get("students", 0)

        group_by_row = self._build_group_by_row(workbook)

        for row_index, row_data in self._read_table_rows(workbook, table):
            surname = self._clean_string(row_data.get("surname"))
            first_name = self._clean_string(row_data.get("first_name"))
            group_name = self._clean_string(row_data.get("group_name"))

            if not group_name:
                group_name = group_by_row.get((table.sheet_name, row_index))

            if not surname or not first_name:
                result.warnings.append(
                    f"{table.sheet_name}, row {row_index}: student surname/first_name is empty, skipped."
                )
                continue

            if not group_name:
                result.errors.append(
                    f"{table.sheet_name}, row {row_index}: cannot detect student group."
                )
                continue

            group_id = self.group_name_to_id.get(group_name)

            if group_id is None:
                result.errors.append(
                    f"{table.sheet_name}, row {row_index}: group {group_name!r} was not imported."
                )
                continue

            student = self._get_or_create_student(
                group_id=group_id,
                surname=surname,
                first_name=first_name,
                bmstu_email=self._clean_string(row_data.get("bmstu_email")),
            )

            full_name_key = self._student_key(surname, first_name)
            surname_key = surname.strip().lower()

            self.student_name_to_id[full_name_key] = student.id
            self.student_surname_to_ids.setdefault(surname_key, [])

            if student.id not in self.student_surname_to_ids[surname_key]:
                self.student_surname_to_ids[surname_key].append(student.id)

            if student.created_at == student.updated_at:
                created_count += 1

        result.created["students"] = created_count

    def _import_marks(
        self,
        workbook: Any,
        table: FoundTable,
        result: HumanTableImportResult,
    ) -> None:
        created_count = result.created.get("marks", 0)

        lesson = self._get_or_create_default_lesson()
        self._link_all_groups_to_schedule(lesson.schedule_id)

        for row_index, row_data in self._read_table_rows(workbook, table):
            student_text = self._clean_string(row_data.get("student"))
            mark_value = row_data.get("mark")

            if not student_text or mark_value is None:
                result.warnings.append(
                    f"{table.sheet_name}, row {row_index}: student or mark is empty, skipped."
                )
                continue

            student_id = self._resolve_student_id(student_text)

            if student_id is None:
                result.errors.append(
                    f"{table.sheet_name}, row {row_index}: cannot resolve student {student_text!r}."
                )
                continue

            try:
                mark_int = int(mark_value)
            except (TypeError, ValueError):
                result.errors.append(
                    f"{table.sheet_name}, row {row_index}: mark {mark_value!r} is not an integer."
                )
                continue

            existing_mark = self.session.scalar(
                select(Mark).where(
                    Mark.student_id == student_id,
                    Mark.lesson_id == lesson.id,
                )
            )

            if existing_mark is not None:
                existing_mark.data = mark_int
                result.warnings.append(
                    f"{table.sheet_name}, row {row_index}: existing mark was updated."
                )
                continue

            mark = Mark(
                student_id=student_id,
                lesson_id=lesson.id,
                data=mark_int,
            )
            self.session.add(mark)
            self.session.flush()

            created_count += 1

        result.created["marks"] = created_count

    def _build_group_by_row(self, workbook: Any) -> dict[tuple[str, int], str]:
        """
        Для файлов вроде pohuy.xlsx:
        группа и студент лежат на одной строке, но в разных таблицах.

        Например:
        C5:D5 -> группа
        G5:I5 -> студент

        Тогда студенту с G5 назначаем группу из C5.
        """
        group_by_row: dict[tuple[str, int], str] = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row_index in range(1, sheet.max_row + 1):
                for col_index in range(1, sheet.max_column + 1):
                    value = sheet.cell(row=row_index, column=col_index).value

                    if value is None:
                        continue

                    value_text = str(value).strip()

                    if value_text in self.group_name_to_id:
                        group_by_row[(sheet_name, row_index)] = value_text

        return group_by_row

    def _get_or_create_group(self, name: str, speciality: str | None) -> Group:
        group = self.session.scalar(
            select(Group).where(Group.name == name)
        )

        if group is not None:
            if speciality is not None and group.speciality != speciality:
                group.speciality = speciality
                self.session.flush()

            return group

        group = Group(name=name, speciality=speciality)
        self.session.add(group)
        self.session.flush()

        return group

    def _get_or_create_student(
            self,
            *,
            group_id: int,
            surname: str,
            first_name: str,
            bmstu_email: str | None,
    ) -> Student:
        student = self.session.scalar(
            select(Student).where(
                Student.group_id == group_id,
                Student.surname == surname,
                Student.first_name == first_name,
            )
        )

        if student is not None:
            if bmstu_email is not None and student.bmstu_email != bmstu_email:
                student.bmstu_email = bmstu_email
                self.session.flush()

            return student

        student = Student(
            group_id=group_id,
            surname=surname,
            first_name=first_name,
            bmstu_email=bmstu_email,
        )
        self.session.add(student)
        self.session.flush()

        return student

    def _get_or_create_default_lesson(self) -> Lesson:
        schedule = self.session.scalar(
            select(Schedule).where(
                Schedule.odd_or_even == "all",
                Schedule.type == "human-table-import",
                Schedule.day == "monday",
                Schedule.time == time(hour=9, minute=0),
            )
        )

        if schedule is None:
            schedule = Schedule(
                odd_or_even="all",
                type="human-table-import",
                is_assessment=True,
                day="monday",
                time=time(hour=9, minute=0),
            )
            self.session.add(schedule)
            self.session.flush()

        today = date.today()

        lesson = self.session.scalar(
            select(Lesson).where(
                Lesson.schedule_id == schedule.id,
                Lesson.date == today,
            )
        )

        if lesson is not None:
            return lesson

        lesson = Lesson(
            schedule_id=schedule.id,
            topic="Human table import",
            date=today,
        )
        self.session.add(lesson)
        self.session.flush()

        return lesson

    def _link_all_groups_to_schedule(self, schedule_id: int) -> None:
        for group_id in self.group_name_to_id.values():
            existing_link = self.session.scalar(
                select(ScheduleGroupLink).where(
                    ScheduleGroupLink.group_id == group_id,
                    ScheduleGroupLink.schedule_id == schedule_id,
                )
            )

            if existing_link is not None:
                continue

            link = ScheduleGroupLink(
                group_id=group_id,
                schedule_id=schedule_id,
            )
            self.session.add(link)

        self.session.flush()

    def _resolve_student_id(self, student_text: str) -> int | None:
        normalized = student_text.strip().lower()

        # Полное имя, если вдруг в Excel будет "Бобков Борис".
        parts = normalized.split()

        if len(parts) >= 2:
            full_name_key = self._student_key(parts[0], parts[1])
            student_id = self.student_name_to_id.get(full_name_key)

            if student_id is not None:
                return student_id

        # Для pohuy.xlsx: на Лист2 только "Бобков".
        candidates = self.student_surname_to_ids.get(normalized, [])

        if len(candidates) == 1:
            return candidates[0]

        return None

    @staticmethod
    def _student_key(surname: str, first_name: str) -> str:
        return f"{surname.strip().lower()}::{first_name.strip().lower()}"

    @staticmethod
    def _clean_string(value: Any) -> str | None:
        if value is None:
            return None

        text = str(value).strip()

        if not text:
            return None

        return text


def import_xlsx_with_human_table_import(
    file_path: str | Path,
) -> HumanTableImportResult:
    db_url = settings.get_db_url()

    engine = create_engine(db_url)
    session_factory = sessionmaker(bind=engine)
    session = session_factory()

    try:
        importer = HumanTableImporter(session)
        return importer.import_workbook(file_path)
    finally:
        session.close()
        engine.dispose()