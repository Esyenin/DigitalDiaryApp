"""
Multi-sheet smart import для XLSX-файлов дневника.

Этот импорт нужен для файлов, где данные разложены по нескольким листам:

groups
schedules
students
schedule_group_links
lessons
attendances
marks
comments

Главное отличие от обычного smart import:
он сохраняет соответствие старых id из Excel новым id из базы данных
и автоматически пересчитывает внешние ключи.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

from config import settings
from app.io_tools.xlsx_config import XLSX_SHEETS_ORDER
from app.io_tools.tabular.entity_schema_rules import CREATE_SCHEMA_BY_ENTITY
from app.models import (
    Attendance,
    Comment,
    Group,
    Lesson,
    Mark,
    Schedule,
    ScheduleGroupLink,
    Student,
)


SERVICE_FIELDS = {"id", "created_at", "updated_at"}
SHEET_ALIASES_BY_ENTITY = {
    "groups": {"groups", "Groups Data", "Group Data", "Группы"},
    "students": {"students", "Students Data", "Student Data", "Студенты"},
    "schedules": {"schedules", "Schedules Data", "Расписание"},
    "schedule_group_links": {
        "schedule_group_links",
        "Schedule Group Links",
        "Schedule Group Links Data",
    },
    "lessons": {"lessons", "Lessons Data", "Уроки", "Занятия"},
    "attendances": {"attendances", "Attendances Data", "Посещаемость"},
    "marks": {"marks", "Marks Data", "Оценки"},
    "comments": {"comments", "Comments Data", "Комментарии"},
}
COLUMN_ALIASES_BY_ENTITY = {
    "groups": {
        "Group": "name",
        "group": "name",
        "Группа": "name",
        "Name": "name",
        "name": "name",

        "Speciality": "speciality",
        "speciality": "speciality",
        "Специальность": "speciality",
    },

    "students": {
        "Last Name": "surname",
        "Surname": "surname",
        "Фамилия": "surname",

        "First Name": "first_name",
        "Name": "first_name",
        "Имя": "first_name",

        "Group": "group_name",
        "group": "group_name",
        "Группа": "group_name",

        "Email": "bmstu_email",
        "email": "bmstu_email",
        "Почта": "bmstu_email",
    },
}
MODEL_BY_ENTITY = {
    "groups": Group,
    "schedules": Schedule,
    "students": Student,
    "schedule_group_links": ScheduleGroupLink,
    "lessons": Lesson,
    "attendances": Attendance,
    "marks": Mark,
    "comments": Comment,
}

FOREIGN_KEYS_BY_ENTITY = {
    "students": {
        "group_id": "groups",
    },
    "schedule_group_links": {
        "group_id": "groups",
        "schedule_id": "schedules",
    },
    "lessons": {
        "schedule_id": "schedules",
    },
    "attendances": {
        "student_id": "students",
        "lesson_id": "lessons",
    },
    "marks": {
        "student_id": "students",
        "lesson_id": "lessons",
    },
    "comments": {
        "student_id": "students",
        "lesson_id": "lessons",
    },
}


@dataclass(slots=True)
class MultiSheetSmartImportResult:
    created: dict[str, int] = field(default_factory=dict)
    old_to_new_ids: dict[str, dict[int, int]] = field(default_factory=dict)
    group_name_to_id: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not self.errors


class MultiSheetSmartImporter:
    """
    Импортирует XLSX-книгу целиком, проходя по листам в правильном порядке.
    """
    def _get_or_create_group_from_payload(
        self,
        payload: dict[str, Any],
    ) -> Group:
        name = str(payload["name"]).strip()
        speciality = payload.get("speciality")

        group = self.session.scalar(
            select(Group).where(Group.name == name)
        )

        if group is not None:
            if speciality is not None and group.speciality != speciality:
                group.speciality = speciality
                self.session.flush()

            return group

        group = Group(
            name=name,
            speciality=speciality,
        )
        self.session.add(group)
        self.session.flush()

        return group

    def _get_or_create_student_from_payload(
        self,
        payload: dict[str, Any],
    ) -> Student:
        student = self.session.scalar(
            select(Student).where(
                Student.group_id == payload["group_id"],
                Student.surname == payload["surname"],
                Student.first_name == payload["first_name"],
            )
        )

        if student is not None:
            for field_name in (
                "patronymic",
                "personal_data",
                "bmstu_email",
            ):
                if field_name in payload and getattr(student, field_name) != payload[field_name]:
                    setattr(student, field_name, payload[field_name])

            self.session.flush()
            return student

        student = Student(**payload)
        self.session.add(student)
        self.session.flush()

        return student

    def _create_or_update_instance(
            self,
            entity_type: str,
            payload: dict[str, Any],
    ) -> Any:
        if entity_type == "groups":
            return self._get_or_create_group_from_payload(payload)

        if entity_type == "students":
            return self._get_or_create_student_from_payload(payload)

        if entity_type == "schedules":
            schedule = self.session.scalar(
                select(Schedule).where(
                    Schedule.day == payload["day"],
                    Schedule.time == payload["time"],
                    Schedule.odd_or_even == payload["odd_or_even"],
                    Schedule.type == payload["type"],
                )
            )

            if schedule is not None:
                if "is_assessment" in payload:
                    schedule.is_assessment = payload["is_assessment"]
                self.session.flush()
                return schedule

            schedule = Schedule(**payload)
            self.session.add(schedule)
            self.session.flush()
            return schedule

        if entity_type == "schedule_group_links":
            link = self.session.scalar(
                select(ScheduleGroupLink).where(
                    ScheduleGroupLink.group_id == payload["group_id"],
                    ScheduleGroupLink.schedule_id == payload["schedule_id"],
                )
            )

            if link is not None:
                return link

            link = ScheduleGroupLink(**payload)
            self.session.add(link)
            self.session.flush()
            return link

        if entity_type == "lessons":
            lesson = self.session.scalar(
                select(Lesson).where(
                    Lesson.schedule_id == payload["schedule_id"],
                    Lesson.date == payload["date"],
                )
            )

            if lesson is not None:
                if "topic" in payload:
                    lesson.topic = payload["topic"]
                self.session.flush()
                return lesson

            lesson = Lesson(**payload)
            self.session.add(lesson)
            self.session.flush()
            return lesson

        if entity_type == "attendances":
            attendance = self.session.scalar(
                select(Attendance).where(
                    Attendance.student_id == payload["student_id"],
                    Attendance.lesson_id == payload["lesson_id"],
                )
            )

            if attendance is not None:
                if "is_visited" in payload:
                    attendance.is_visited = payload["is_visited"]
                self.session.flush()
                return attendance

            attendance = Attendance(**payload)
            self.session.add(attendance)
            self.session.flush()
            return attendance

        if entity_type == "marks":
            mark = self.session.scalar(
                select(Mark).where(
                    Mark.student_id == payload["student_id"],
                    Mark.lesson_id == payload["lesson_id"],
                )
            )

            if mark is not None:
                if "data" in payload:
                    mark.data = payload["data"]
                self.session.flush()
                return mark

            mark = Mark(**payload)
            self.session.add(mark)
            self.session.flush()
            return mark

        if entity_type == "comments":
            comment = self.session.scalar(
                select(Comment).where(
                    Comment.student_id == payload["student_id"],
                    Comment.lesson_id == payload["lesson_id"],
                )
            )

            if comment is not None:
                if "data" in payload:
                    comment.data = payload["data"]
                self.session.flush()
                return comment

            comment = Comment(**payload)
            self.session.add(comment)
            self.session.flush()
            return comment

        model = MODEL_BY_ENTITY[entity_type]
        instance = model(**payload)
        self.session.add(instance)
        self.session.flush()
        return instance
    @staticmethod
    def _find_header_row_index(
            rows: list[tuple[Any, ...]],
            entity_type: str,
    ) -> int | None:
        supported_headers_by_entity: dict[str, set[str]] = {
            "groups": {
                "id",
                "name",
                "group",
                "speciality",
                "created_at",
                "updated_at",
                "группа",
                "специальность",
            },
            "schedules": {
                "id",
                "odd_or_even",
                "type",
                "is_assessment",
                "day",
                "time",
                "created_at",
                "updated_at",
                "четность",
                "тип",
                "день",
                "время",
            },
            "students": {
                "id",
                "group_id",
                "surname",
                "first_name",
                "patronymic",
                "personal_data",
                "bmstu_email",
                "created_at",
                "updated_at",
                "last name",
                "first name",
                "group",
                "email",
                "фамилия",
                "имя",
                "отчество",
                "группа",
                "почта",
            },
            "schedule_group_links": {
                "id",
                "group_id",
                "schedule_id",
                "created_at",
                "updated_at",
            },
            "lessons": {
                "id",
                "schedule_id",
                "topic",
                "date",
                "created_at",
                "updated_at",
                "тема",
                "дата",
            },
            "attendances": {
                "id",
                "student_id",
                "lesson_id",
                "is_visited",
                "created_at",
                "updated_at",
                "посещение",
                "присутствовал",
            },
            "marks": {
                "id",
                "student_id",
                "lesson_id",
                "data",
                "created_at",
                "updated_at",
                "mark",
                "grade",
                "оценка",
                "балл",
            },
            "comments": {
                "id",
                "student_id",
                "lesson_id",
                "data",
                "created_at",
                "updated_at",
                "comment",
                "comments",
                "комментарий",
            },
        }

        supported_headers = supported_headers_by_entity.get(entity_type, set())

        for index, row in enumerate(rows):
            normalized_values = {
                str(value).strip().lower()
                for value in row
                if value is not None and str(value).strip()
            }

            if normalized_values & supported_headers:
                return index

        return None
    @staticmethod
    def _find_sheet_name(workbook: Any, entity_type: str) -> str | None:
        aliases = SHEET_ALIASES_BY_ENTITY.get(entity_type, {entity_type})

        for sheet_name in workbook.sheetnames:
            if sheet_name in aliases:
                return sheet_name

        return None
    def __init__(self, session: Session) -> None:
        self.session = session

    def import_workbook(self, file_path: str | Path) -> MultiSheetSmartImportResult:
        path = Path(file_path)
        workbook = load_workbook(path, data_only=True)

        result = MultiSheetSmartImportResult()

        found_supported_sheets = []

        for entity_type in XLSX_SHEETS_ORDER:
            sheet_name = self._find_sheet_name(workbook, entity_type)
            if sheet_name is not None:
                found_supported_sheets.append(sheet_name)

        if not found_supported_sheets:
            result.errors.append(
                "В книге не найдено ни одного поддерживаемого листа для "
                "multi-sheet smart import.\n\n"
                f"Найдены листы: {', '.join(workbook.sheetnames)}"
            )
            return result

        for entity_type in XLSX_SHEETS_ORDER:
            sheet_name = self._find_sheet_name(workbook, entity_type)

            if sheet_name is None:
                result.warnings.append(
                    f"Sheet for {entity_type!r} was skipped because it does not exist."
                )
                continue

            self._import_sheet(
                workbook=workbook,
                sheet_name=sheet_name,
                entity_type=entity_type,
                result=result,
            )

            if result.errors:
                self.session.rollback()
                return result

        self.session.commit()
        return result

    def _import_sheet(
            self,
            *,
            workbook: Any,
            sheet_name: str,
            entity_type: str,
            result: MultiSheetSmartImportResult,
    ) -> None:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            result.warnings.append(f"Sheet {entity_type!r} is empty.")
            return

        header_row_index = self._find_header_row_index(rows, entity_type)

        if header_row_index is None:
            result.warnings.append(
                f"Sheet {sheet_name!r} was skipped because no supported header row was found."
            )
            return

        headers = [
            self._normalize_header(entity_type, value)
            for value in rows[header_row_index]
        ]

        data_rows = rows[header_row_index + 1:]

        if not any(headers):
            result.warnings.append(f"Sheet {sheet_name!r} has empty header row.")
            return

        created_count = 0

        for excel_row_number, row_values in enumerate(
                data_rows,
                start=header_row_index + 2,
        ):
            raw_row = self._row_to_dict(headers, row_values)

            if not raw_row:
                continue

            old_id = self._to_int_or_none(raw_row.get("id"))

            try:
                payload = self._prepare_payload(
                    entity_type=entity_type,
                    raw_row=raw_row,
                    result=result,
                    excel_row_number=excel_row_number,
                )
            except ValueError as exc:
                result.errors.append(str(exc))
                continue

            if payload is None:
                continue

            try:
                instance = self._create_or_update_instance(
                    entity_type=entity_type,
                    payload=payload,
                )
            except Exception as exc:
                result.errors.append(
                    f"{entity_type}, row {excel_row_number}: database error: {exc}"
                )
                continue

            new_id = getattr(instance, "id", None)
            if entity_type == "groups" and new_id is not None:
                group_name = payload.get("name")
                if group_name is not None:
                    result.group_name_to_id[str(group_name).strip()] = int(new_id)

            if old_id is not None and new_id is not None:
                result.old_to_new_ids.setdefault(entity_type, {})[old_id] = int(new_id)

            created_count += 1

        result.created[entity_type] = created_count

    def _prepare_payload(
        self,
        *,
        entity_type: str,
        raw_row: dict[str, Any],
        result: MultiSheetSmartImportResult,
        excel_row_number: int,
    ) -> dict[str, Any] | None:
        payload = {
            key: value
            for key, value in raw_row.items()
            if key not in SERVICE_FIELDS and value is not None
        }

        resolved_group_by_name = False

        if entity_type == "students" and "group_name" in payload:
            group_name = str(payload.pop("group_name")).strip()
            group_id = result.group_name_to_id.get(group_name)

            if group_id is None:
                raise ValueError(
                    f"students, row {excel_row_number}: cannot find group "
                    f"with name {group_name!r}. Import groups first."
                )

            payload["group_id"] = group_id
            resolved_group_by_name = True

        if not resolved_group_by_name:
            self._remap_foreign_keys(
                entity_type=entity_type,
                payload=payload,
                result=result,
                excel_row_number=excel_row_number,
            )

        payload = self._normalize_special_values(entity_type, payload)

        schema = CREATE_SCHEMA_BY_ENTITY.get(entity_type)
        if schema is None:
            raise ValueError(
                f"{entity_type}, row {excel_row_number}: create schema not found."
            )

        try:
            validated = schema.model_validate(payload)
        except ValidationError as exc:
            messages = "; ".join(
                f"{'.'.join(str(part) for part in error['loc'])}: {error['msg']}"
                for error in exc.errors()
            )
            raise ValueError(
                f"{entity_type}, row {excel_row_number}: validation error: {messages}"
            ) from exc

        return validated.model_dump(exclude_unset=True)

    def _remap_foreign_keys(
        self,
        *,
        entity_type: str,
        payload: dict[str, Any],
        result: MultiSheetSmartImportResult,
        excel_row_number: int,
    ) -> None:
        foreign_keys = FOREIGN_KEYS_BY_ENTITY.get(entity_type, {})

        for field_name, referenced_entity in foreign_keys.items():
            old_value = self._to_int_or_none(payload.get(field_name))

            if old_value is None:
                continue

            new_value = result.old_to_new_ids.get(referenced_entity, {}).get(old_value)

            if new_value is None:
                raise ValueError(
                    f"{entity_type}, row {excel_row_number}: cannot remap "
                    f"{field_name}={old_value}. No imported {referenced_entity} "
                    f"record with old id {old_value}."
                )

            payload[field_name] = new_value

    @staticmethod
    def _normalize_header(entity_type: str, value: Any) -> str:
        if value is None:
            return ""

        header = str(value).strip()

        entity_aliases = COLUMN_ALIASES_BY_ENTITY.get(entity_type, {})
        return entity_aliases.get(header, header)

    @staticmethod
    def _row_to_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
        row: dict[str, Any] = {}

        for header, value in zip(headers, values):
            if not header:
                continue

            if value == "":
                value = None

            row[header] = value

        return {
            key: value
            for key, value in row.items()
            if value is not None
        }

    @staticmethod
    def _to_int_or_none(value: Any) -> int | None:
        if value is None:
            return None

        if isinstance(value, bool):
            return None

        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _normalize_special_values(
        entity_type: str,
        payload: dict[str, Any],
    ) -> dict[str, Any]:
        normalized = dict(payload)

        if entity_type == "schedules":
            value = normalized.get("time")
            if isinstance(value, float | int):
                normalized["time"] = MultiSheetSmartImporter._excel_fraction_to_time(
                    value
                )

        if entity_type == "lessons":
            value = normalized.get("date")
            if isinstance(value, datetime):
                normalized["date"] = value.date()
            elif isinstance(value, float | int):
                normalized["date"] = MultiSheetSmartImporter._excel_serial_to_date(
                    value
                )

        return normalized

    @staticmethod
    def _excel_serial_to_date(value: float | int) -> date:
        """
        Конвертирует Excel serial date в date.

        Для XLSX обычно используется база 1899-12-30.
        """
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()

    @staticmethod
    def _excel_fraction_to_time(value: float | int) -> time:
        """
        Конвертирует Excel time fraction в time.

        Например:
        0.5 -> 12:00:00
        """
        total_seconds = round(float(value) * 24 * 60 * 60)
        total_seconds = total_seconds % (24 * 60 * 60)

        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60

        return time(hour=hours, minute=minutes, second=seconds)


def import_xlsx_with_multisheet_smart_import(
    file_path: str | Path,
) -> MultiSheetSmartImportResult:
    db_url = settings.get_db_url()

    engine = create_engine(db_url)
    session_factory = sessionmaker(bind=engine)
    session = session_factory()

    try:
        importer = MultiSheetSmartImporter(session)
        return importer.import_workbook(file_path)
    finally:
        session.close()
        engine.dispose()