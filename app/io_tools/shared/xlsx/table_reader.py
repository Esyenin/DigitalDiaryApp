"""
Канонический XLSX-reader подсистемы `io_tools`.

В модуле сосредоточено всё активное чтение XLSX:

1. поиск таблиц в неструктурированных книгах;
2. чтение произвольного диапазона;
3. чтение стандартной XLSX-книги приложения;
4. фасад `XlsxTableReader` для flow-сценариев.

Старые модули `xlsx_importer/raw_reader.py` и `formats/xlsx/readers/*`
должны использоваться только как compatibility wrappers к этим классам.
"""
from __future__ import annotations

from dataclasses import asdict
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.io_tools.shared.results import TabularImportResult
from app.io_tools.shared.tabular.models import ExtractedTable, TableRegion
from app.io_tools.shared.tabular.schema_registry import classify_tabular_headers
from app.io_tools.shared.xlsx.config import (
    XLSX_COLUMNS_BY_SHEET,
    XLSX_REQUIRED_COLUMNS_BY_SHEET,
    XLSX_SHEETS_ORDER,
    normalize_sheet_keys,
)


logger = logging.getLogger(__name__)


class RawWorkbookReader:
    """
    Ищет табличные области в листах Excel-книги.
    """

    def __init__(
        self,
        *,
        min_cells_in_row: int = 2,
        min_cells_in_col: int = 2,
        max_row_gap: int = 1,
        max_col_gap: int = 1,
        min_rows: int = 2,
        min_cols: int = 2,
    ) -> None:
        self.min_cells_in_row = min_cells_in_row
        self.min_cells_in_col = min_cells_in_col
        self.max_row_gap = max_row_gap
        self.max_col_gap = max_col_gap
        self.min_rows = min_rows
        self.min_cols = min_cols

    def find_tables_in_workbook(
        self,
        path: str | Path,
        min_score: float = 0.45,
    ) -> list[TableRegion]:
        logger.info(
            "RawWorkbookReader workbook scan started. source=%s min_score=%s.",
            path,
            min_score,
        )
        workbook = load_workbook(path, data_only=True)
        regions: list[TableRegion] = []
        for worksheet in workbook.worksheets:
            regions.extend(self.find_tables_on_sheet(worksheet, min_score=min_score))
        logger.info(
            "RawWorkbookReader workbook scan finished. source=%s tables_count=%s.",
            path,
            len(regions),
        )
        return regions

    def find_tables_on_sheet(
        self,
        worksheet: Worksheet,
        min_score: float = 0.45,
    ) -> list[TableRegion]:
        raw_regions = self._build_candidate_regions(worksheet)
        tables: list[TableRegion] = []
        for region in raw_regions:
            trimmed_region = self._trim_weak_edges(worksheet, region)
            if not self._is_likely_table_region(
                worksheet,
                trimmed_region,
                min_score=min_score,
            ):
                continue

            stats = self._region_stats(worksheet, trimmed_region)
            score = self._table_likeness_score(worksheet, trimmed_region)
            tables.append(
                TableRegion(
                    sheet=worksheet.title,
                    range=self._region_to_excel_range(trimmed_region),
                    min_row=trimmed_region["min_row"],
                    max_row=trimmed_region["max_row"],
                    min_col=trimmed_region["min_col"],
                    max_col=trimmed_region["max_col"],
                    rows=int(stats["rows"]),
                    cols=int(stats["cols"]),
                    total_cells=int(stats["total_cells"]),
                    non_empty_cells=int(stats["non_empty_cells"]),
                    density=float(stats["density"]),
                    score=score,
                )
            )
        return tables

    @staticmethod
    def normalize_cell_value(value: Any) -> Any | None:
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            if value == "":
                return None
        return value

    @classmethod
    def is_non_empty(cls, value: Any) -> bool:
        return cls.normalize_cell_value(value) is not None

    def _row_non_empty_counts(self, worksheet: Worksheet) -> dict[int, int]:
        counts: dict[int, int] = {}
        for row_index in range(worksheet.min_row, worksheet.max_row + 1):
            counts[row_index] = sum(
                1
                for col_index in range(worksheet.min_column, worksheet.max_column + 1)
                if self.is_non_empty(worksheet.cell(row_index, col_index).value)
            )
        return counts

    def _column_non_empty_counts(
        self,
        worksheet: Worksheet,
        row_start: int,
        row_end: int,
    ) -> dict[int, int]:
        counts: dict[int, int] = {}
        for col_index in range(worksheet.min_column, worksheet.max_column + 1):
            counts[col_index] = sum(
                1
                for row_index in range(row_start, row_end + 1)
                if self.is_non_empty(worksheet.cell(row_index, col_index).value)
            )
        return counts

    @staticmethod
    def _find_bands(
        counts: dict[int, int],
        *,
        min_count: int,
        max_gap: int,
        min_size: int,
    ) -> list[tuple[int, int]]:
        bands: list[tuple[int, int]] = []
        current_start: int | None = None
        current_end: int | None = None
        gap_count = 0

        for index in sorted(counts):
            looks_useful = counts[index] >= min_count
            if looks_useful:
                if current_start is None:
                    current_start = index
                current_end = index
                gap_count = 0
                continue

            if current_start is None:
                continue

            gap_count += 1
            if gap_count > max_gap:
                assert current_end is not None
                if current_end - current_start + 1 >= min_size:
                    bands.append((current_start, current_end))
                current_start = None
                current_end = None
                gap_count = 0

        if current_start is not None and current_end is not None:
            if current_end - current_start + 1 >= min_size:
                bands.append((current_start, current_end))
        return bands

    @staticmethod
    def _region_to_excel_range(region: dict[str, int]) -> str:
        start_col = get_column_letter(region["min_col"])
        end_col = get_column_letter(region["max_col"])
        return f"{start_col}{region['min_row']}:{end_col}{region['max_row']}"

    def _region_stats(
        self,
        worksheet: Worksheet,
        region: dict[str, int],
    ) -> dict[str, int | float]:
        total_cells = 0
        non_empty_cells = 0
        for row_index in range(region["min_row"], region["max_row"] + 1):
            for col_index in range(region["min_col"], region["max_col"] + 1):
                total_cells += 1
                if self.is_non_empty(worksheet.cell(row_index, col_index).value):
                    non_empty_cells += 1

        rows = region["max_row"] - region["min_row"] + 1
        cols = region["max_col"] - region["min_col"] + 1
        return {
            "rows": rows,
            "cols": cols,
            "total_cells": total_cells,
            "non_empty_cells": non_empty_cells,
            "density": non_empty_cells / total_cells if total_cells else 0.0,
        }

    def _trim_weak_edges(
        self,
        worksheet: Worksheet,
        region: dict[str, int],
        min_filled_ratio: float = 0.3,
    ) -> dict[str, int]:
        min_row = region["min_row"]
        max_row = region["max_row"]
        min_col = region["min_col"]
        max_col = region["max_col"]

        def row_ratio(row_index: int) -> float:
            total = max_col - min_col + 1
            filled = sum(
                1
                for col_index in range(min_col, max_col + 1)
                if self.is_non_empty(worksheet.cell(row_index, col_index).value)
            )
            return filled / total if total else 0.0

        def col_ratio(col_index: int) -> float:
            total = max_row - min_row + 1
            filled = sum(
                1
                for row_index in range(min_row, max_row + 1)
                if self.is_non_empty(worksheet.cell(row_index, col_index).value)
            )
            return filled / total if total else 0.0

        while min_row <= max_row and row_ratio(min_row) < min_filled_ratio:
            min_row += 1
        while max_row >= min_row and row_ratio(max_row) < min_filled_ratio:
            max_row -= 1
        while min_col <= max_col and col_ratio(min_col) < min_filled_ratio:
            min_col += 1
        while max_col >= min_col and col_ratio(max_col) < min_filled_ratio:
            max_col -= 1

        return {
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
        }

    def _table_likeness_score(
        self,
        worksheet: Worksheet,
        region: dict[str, int],
    ) -> float:
        stats = self._region_stats(worksheet, region)
        rows = int(stats["rows"])
        cols = int(stats["cols"])
        density = float(stats["density"])
        if rows <= 1 or cols <= 1:
            return 0.0

        header_row = region["min_row"]
        header_values = [
            self.normalize_cell_value(worksheet.cell(header_row, col_index).value)
            for col_index in range(region["min_col"], region["max_col"] + 1)
        ]
        filled_headers = sum(value is not None for value in header_values)
        unique_headers = len(
            {
                str(value).strip().lower()
                for value in header_values
                if value is not None
            }
        )
        header_score = filled_headers / cols if cols else 0.0
        uniqueness_score = unique_headers / filled_headers if filled_headers else 0.0
        body_rows = max(rows - 1, 1)
        body_non_empty = max(int(stats["non_empty_cells"]) - filled_headers, 0)
        body_density = body_non_empty / (body_rows * cols) if body_rows * cols else 0.0

        return round(
            density * 0.35
            + header_score * 0.25
            + uniqueness_score * 0.2
            + body_density * 0.2,
            4,
        )

    def _is_likely_table_region(
        self,
        worksheet: Worksheet,
        region: dict[str, int],
        *,
        min_score: float,
    ) -> bool:
        stats = self._region_stats(worksheet, region)
        if int(stats["rows"]) < self.min_rows or int(stats["cols"]) < self.min_cols:
            return False
        if float(stats["density"]) < 0.35:
            return False
        return self._table_likeness_score(worksheet, region) >= min_score

    def _build_candidate_regions(self, worksheet: Worksheet) -> list[dict[str, int]]:
        row_counts = self._row_non_empty_counts(worksheet)
        row_bands = self._find_bands(
            row_counts,
            min_count=self.min_cells_in_row,
            max_gap=self.max_row_gap,
            min_size=self.min_rows,
        )
        regions: list[dict[str, int]] = []
        for row_start, row_end in row_bands:
            col_counts = self._column_non_empty_counts(
                worksheet,
                row_start=row_start,
                row_end=row_end,
            )
            col_bands = self._find_bands(
                col_counts,
                min_count=self.min_cells_in_col,
                max_gap=self.max_col_gap,
                min_size=self.min_cols,
            )
            for col_start, col_end in col_bands:
                regions.append(
                    {
                        "min_row": row_start,
                        "max_row": row_end,
                        "min_col": col_start,
                        "max_col": col_end,
                    }
                )
        return regions


class XlsxRangeReader:
    """
    Читает конкретный диапазон листа как таблицу.
    """

    def read_range(
        self,
        file_path: str | Path,
        sheet_name: str,
        cell_range: str,
        *,
        entity_type: str | None = None,
    ) -> ExtractedTable:
        logger.info(
            "XlsxRangeReader range read started. source=%s sheet=%s range=%s entity_type=%s.",
            file_path,
            sheet_name,
            cell_range,
            entity_type,
        )
        workbook = load_workbook(file_path, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name} was not found in workbook.")

        worksheet = workbook[sheet_name]
        extracted_table = self.read_range_from_worksheet(
            worksheet,
            cell_range,
            entity_type=entity_type,
        )
        logger.info(
            "XlsxRangeReader range read finished. sheet=%s range=%s rows=%s errors=%s.",
            extracted_table.sheet,
            extracted_table.range,
            len(extracted_table.rows),
            len(extracted_table.errors),
        )
        return extracted_table

    def read_detected_table(
        self,
        file_path: str | Path,
        table_region: TableRegion,
        *,
        entity_type: str | None = None,
    ) -> ExtractedTable:
        return self.read_range(
            file_path,
            table_region.sheet,
            table_region.range,
            entity_type=entity_type,
        )

    def read_range_from_worksheet(
        self,
        worksheet: Worksheet,
        cell_range: str,
        *,
        entity_type: str | None = None,
    ) -> ExtractedTable:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        raw_rows = list(
            worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        )

        if not raw_rows:
            raise ValueError(f"Range {cell_range} does not contain rows.")

        headers, warnings = self._build_headers(raw_rows[0])
        if not headers:
            raise ValueError(f"Range {cell_range} does not contain headers.")

        rows = self._build_rows(headers, raw_rows[1:])
        canonical_entity_type = self._normalize_entity_type(entity_type)
        known_headers, unknown_headers, missing_required_headers = self._classify_headers(
            canonical_entity_type,
            headers,
        )

        errors: list[str] = []
        if canonical_entity_type and missing_required_headers:
            errors.append(
                "Missing required headers: "
                + ", ".join(missing_required_headers)
                + "."
            )
        if not rows:
            errors.append("Selected range does not contain data rows.")

        return ExtractedTable(
            sheet=worksheet.title,
            range=cell_range,
            entity_type=canonical_entity_type,
            headers=headers,
            rows=rows,
            known_headers=known_headers,
            unknown_headers=unknown_headers,
            missing_required_headers=missing_required_headers,
            warnings=warnings,
            errors=errors,
        )

    def _normalize_entity_type(self, entity_type: str | None) -> str | None:
        if entity_type is None:
            return None
        return normalize_sheet_keys([entity_type])[0]

    def _build_headers(self, raw_headers: tuple[Any, ...]) -> tuple[tuple[str, ...], list[str]]:
        headers: list[str] = []
        warnings: list[str] = []
        duplicates: set[str] = set()

        for index, raw_header in enumerate(raw_headers, start=1):
            normalized_header = RawWorkbookReader.normalize_cell_value(raw_header)
            if normalized_header is None:
                generated_header = f"__column_{index}"
                headers.append(generated_header)
                warnings.append(
                    f"Empty header in column {index} was replaced with {generated_header}."
                )
                continue

            header = str(normalized_header).strip()
            if header in headers:
                duplicates.add(header)
                header = f"{header}__{index}"
                warnings.append(f"Duplicate header was renamed to {header}.")
            headers.append(header)

        for duplicated_header in sorted(duplicates):
            warnings.append(
                f"Header {duplicated_header} appeared more than once in the selected range."
            )
        return tuple(headers), warnings

    @staticmethod
    def _build_rows(
        headers: tuple[str, ...],
        raw_rows: list[tuple[Any, ...]],
    ) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for raw_row in raw_rows:
            row = {
                header: raw_row[index] if index < len(raw_row) else None
                for index, header in enumerate(headers)
            }
            if all(value is None for value in row.values()):
                continue
            rows.append(row)
        return rows

    @staticmethod
    def _classify_headers(
        entity_type: str | None,
        headers: tuple[str, ...],
    ) -> tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
        classification = classify_tabular_headers(entity_type, headers)
        return (
            classification.known_headers,
            classification.unknown_headers,
            classification.missing_required_headers,
        )


class SelectedRangeReader:
    """
    Простая совместимая обёртка для чтения выбранного диапазона.
    """

    def __init__(self, range_reader: XlsxRangeReader | None = None) -> None:
        self.range_reader = range_reader or XlsxRangeReader()

    def read(
        self,
        file_path: str | Path,
        sheet_name: str,
        cell_range: str,
        *,
        entity_type: str | None = None,
    ) -> ExtractedTable:
        return self.range_reader.read_range(
            file_path,
            sheet_name,
            cell_range,
            entity_type=entity_type,
        )


class TableRegionFinder:
    """
    Адаптер поиска табличных областей в XLSX-книге.
    """

    def __init__(self, raw_reader: RawWorkbookReader | None = None) -> None:
        self.raw_reader = raw_reader or RawWorkbookReader()

    def find(
        self,
        file_path: str | Path,
        *,
        min_score: float = 0.45,
    ) -> list[TableRegion]:
        return self.raw_reader.find_tables_in_workbook(file_path, min_score=min_score)


class StandardWorkbookReader:
    """
    Читает XLSX-книгу стандартного внутреннего формата приложения.
    """

    def read(self, file_path: str | Path) -> TabularImportResult:
        logger.info("StandardWorkbookReader read started. source=%s.", file_path)
        workbook = load_workbook(file_path)
        result = TabularImportResult()
        sheet_names = workbook.sheetnames

        self._validate_sheet_names(sheet_names, result)
        if not result.is_valid:
            return result

        for sheet_name in self._ordered_sheet_names(sheet_names):
            worksheet = workbook[sheet_name]
            result.data[sheet_name] = self._read_sheet(
                sheet_name,
                worksheet,
                result,
            )

        logger.info(
            "StandardWorkbookReader read finished. source=%s sheets_count=%s errors=%s.",
            file_path,
            len(result.data),
            len(result.errors),
        )
        return result

    def _validate_sheet_names(
        self,
        sheet_names: list[str],
        result: TabularImportResult,
    ) -> None:
        unknown_sheet_names = [
            sheet_name
            for sheet_name in sheet_names
            if sheet_name not in XLSX_COLUMNS_BY_SHEET
        ]
        if unknown_sheet_names:
            joined = ", ".join(unknown_sheet_names)
            result.add_error(f"Неизвестные листы в XLSX-файле: {joined}.")

    @staticmethod
    def _ordered_sheet_names(sheet_names: list[str]) -> list[str]:
        ordered_sheet_names = [
            sheet_name
            for sheet_name in XLSX_SHEETS_ORDER
            if sheet_name in sheet_names
        ]
        ordered_sheet_names.extend(
            sheet_name
            for sheet_name in sheet_names
            if sheet_name not in ordered_sheet_names
        )
        return ordered_sheet_names

    def _read_sheet(
        self,
        sheet_name: str,
        worksheet: Any,
        result: TabularImportResult,
    ) -> list[dict[str, object]]:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            self._validate_headers(sheet_name, (), result)
            return []

        headers = tuple(
            str(header).strip()
            for header in rows[0]
            if header is not None and str(header).strip()
        )
        error_count_before = len(result.errors)
        self._validate_headers(sheet_name, headers, result)
        if len(result.errors) > error_count_before:
            return []

        imported_rows: list[dict[str, object]] = []
        for raw_row in rows[1:]:
            row_data = {
                header: raw_row[index] if index < len(raw_row) else None
                for index, header in enumerate(headers)
            }
            if all(value is None for value in row_data.values()):
                continue
            imported_rows.append(row_data)
        return imported_rows

    def _validate_headers(
        self,
        sheet_name: str,
        headers: tuple[str, ...],
        result: TabularImportResult,
    ) -> None:
        required_headers = XLSX_REQUIRED_COLUMNS_BY_SHEET.get(sheet_name, ())
        missing_headers = [
            header for header in required_headers if header not in headers
        ]
        if missing_headers:
            joined = ", ".join(missing_headers)
            result.add_error(
                f"В листе {sheet_name} отсутствуют обязательные колонки: {joined}.",
                sheet_name=sheet_name,
            )


class XlsxTableReader:
    """
    Предоставляет единый интерфейс чтения XLSX на уровне flow-сценариев.
    """

    def __init__(
        self,
        *,
        raw_workbook_reader: RawWorkbookReader | None = None,
        range_reader: XlsxRangeReader | None = None,
        workbook_reader: StandardWorkbookReader | None = None,
    ) -> None:
        self._raw_workbook_reader = raw_workbook_reader or RawWorkbookReader()
        self._range_reader = range_reader or XlsxRangeReader()
        self._selected_range_reader = SelectedRangeReader(self._range_reader)
        self._table_region_finder = TableRegionFinder(self._raw_workbook_reader)
        self._workbook_reader = workbook_reader or StandardWorkbookReader()

    def read_range(
        self,
        file_path: Path,
        sheet_name: str,
        cell_range: str | None,
        entity_type: str | None = None,
    ) -> ExtractedTable:
        if cell_range is None:
            raise ValueError("cell_range is required for XLSX range reading.")
        return self._selected_range_reader.read(
            file_path,
            sheet_name,
            cell_range,
            entity_type=entity_type,
        )

    def read_workbook(self, file_path: Path) -> TabularImportResult:
        return self._workbook_reader.read(file_path)

    def find_tables(
        self,
        file_path: Path,
        min_score: float = 0.5,
    ) -> list[TableRegion]:
        return self._table_region_finder.find(file_path, min_score=min_score)


def find_tables_on_sheet(
    worksheet: Worksheet,
    min_score: float = 0.45,
) -> list[TableRegion]:
    return RawWorkbookReader().find_tables_on_sheet(worksheet, min_score=min_score)


def find_tables_in_workbook(
    path: str | Path,
    min_score: float = 0.45,
) -> list[TableRegion]:
    return RawWorkbookReader().find_tables_in_workbook(path, min_score=min_score)


def print_tables(tables: list[TableRegion]) -> None:
    if not tables:
        print("Таблицы не найдены.")
        return
    for table in tables:
        print(
            f"{table.sheet} | {table.range} | "
            f"{table.rows}x{table.cols} | "
            f"density={table.density:.2f} | "
            f"score={table.score:.2f}"
        )


def tables_to_dicts(tables: list[TableRegion]) -> list[dict[str, object]]:
    return [asdict(table) for table in tables]
