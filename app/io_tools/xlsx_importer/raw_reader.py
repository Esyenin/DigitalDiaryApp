"""
Низкоуровневое чтение нестандартных XLSX-таблиц.

Модуль решает две базовые задачи:

1. Поиск прямоугольных областей, которые выглядят как таблицы.
2. Чтение конкретного диапазона листа с диагностикой заголовков и строк.

Модуль не занимается импортом в базу и не пытается понять бизнес-смысл
таблицы глубже, чем это нужно для базовой структурной проверки.
"""
from __future__ import annotations

from dataclasses import asdict
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.io_tools.tabular.header_classifier import classify_tabular_headers
from app.io_tools.tabular.models import ExtractedTable, TableRegion
from app.io_tools.xlsx_config import normalize_sheet_keys


logger = logging.getLogger(__name__)


class RawWorkbookReader:
    """
    Ищет табличные области в листах Excel-книги.

    Класс нужен для предварительного анализа нестандартных файлов, в которых
    таблица может начинаться не с первой строки и не с первого столбца.
    """

    def __init__(
        self,
        *,
        min_cells_in_row: int = 2,
        min_cells_in_col: int = 2,
        max_row_gap: int = 1,
        max_col_gap: int = 1,
        min_rows: int = 2,
        min_cols: int = 2,
    ) -> None:
        """
        Создает детектор табличных областей.

        :param min_cells_in_row: Минимум заполненных ячеек в строке, чтобы
            считать ее содержательной для таблицы.
        :param min_cells_in_col: Минимум заполненных ячеек в колонке, чтобы
            считать ее содержательной для таблицы.
        :param max_row_gap: Максимум слабых строк внутри найденной полосы.
        :param max_col_gap: Максимум слабых колонок внутри найденной полосы.
        :param min_rows: Минимальная высота кандидата на таблицу.
        :param min_cols: Минимальная ширина кандидата на таблицу.
        """
        self.min_cells_in_row = min_cells_in_row
        self.min_cells_in_col = min_cells_in_col
        self.max_row_gap = max_row_gap
        self.max_col_gap = max_col_gap
        self.min_rows = min_rows
        self.min_cols = min_cols

    def find_tables_in_workbook(
        self,
        path: str | Path,
        min_score: float = 0.45,
    ) -> list[TableRegion]:
        """
        Ищет таблицы во всех листах Excel-книги.

        :param path: Путь к XLSX-файлу.
        :param min_score: Минимальный итоговый балл похожести на таблицу.
        :return: Список найденных табличных областей во всех листах книги.
        """
        logger.info(
            "RawWorkbookReader workbook scan started. source=%s min_score=%s.",
            path,
            min_score,
        )
        workbook = load_workbook(path, data_only=True)
        regions: list[TableRegion] = []

        for worksheet in workbook.worksheets:
            regions.extend(self.find_tables_on_sheet(worksheet, min_score=min_score))

        logger.info(
            "RawWorkbookReader workbook scan finished. source=%s tables_count=%s.",
            path,
            len(regions),
        )
        return regions

    def find_tables_on_sheet(
        self,
        worksheet: Worksheet,
        min_score: float = 0.45,
    ) -> list[TableRegion]:
        """
        Ищет таблицы на одном листе.

        :param worksheet: Лист Excel для анализа.
        :param min_score: Минимальный итоговый балл похожести на таблицу.
        :return: Список найденных областей таблиц на указанном листе.
        """
        logger.debug(
            "RawWorkbookReader sheet scan started. sheet=%s min_score=%s.",
            worksheet.title,
            min_score,
        )
        raw_regions = self._build_candidate_regions(worksheet)
        tables: list[TableRegion] = []

        for region in raw_regions:
            trimmed_region = self._trim_weak_edges(worksheet, region)

            if not self._is_likely_table_region(
                worksheet,
                trimmed_region,
                min_score=min_score,
            ):
                continue

            stats = self._region_stats(worksheet, trimmed_region)
            score = self._table_likeness_score(worksheet, trimmed_region)
            tables.append(
                TableRegion(
                    sheet=worksheet.title,
                    range=self._region_to_excel_range(trimmed_region),
                    min_row=trimmed_region["min_row"],
                    max_row=trimmed_region["max_row"],
                    min_col=trimmed_region["min_col"],
                    max_col=trimmed_region["max_col"],
                    rows=int(stats["rows"]),
                    cols=int(stats["cols"]),
                    total_cells=int(stats["total_cells"]),
                    non_empty_cells=int(stats["non_empty_cells"]),
                    density=float(stats["density"]),
                    score=score,
                )
            )

        logger.debug(
            "RawWorkbookReader sheet scan finished. sheet=%s tables_count=%s.",
            worksheet.title,
            len(tables),
        )
        return tables

    @staticmethod
    def normalize_cell_value(value: Any) -> Any | None:
        """
        Приводит значение ячейки к удобному виду для анализа.

        :param value: Исходное значение ячейки.
        :return: `None` для фактически пустых значений, иначе очищенное значение.
        """
        if value is None:
            return None

        if isinstance(value, str):
            value = value.strip()
            if value == "":
                return None

        return value

    @classmethod
    def is_non_empty(cls, value: Any) -> bool:
        """
        Проверяет, содержит ли ячейка значимые данные.

        :param value: Значение ячейки.
        :return: `True`, если значение не считается пустым.
        """
        return cls.normalize_cell_value(value) is not None

    def _row_non_empty_counts(self, worksheet: Worksheet) -> dict[int, int]:
        """
        Считает количество непустых ячеек в каждой строке листа.

        :param worksheet: Лист Excel для анализа.
        :return: Словарь вида `номер_строки -> число непустых ячеек`.
        """
        counts: dict[int, int] = {}

        for row_index in range(worksheet.min_row, worksheet.max_row + 1):
            counts[row_index] = sum(
                1
                for col_index in range(worksheet.min_column, worksheet.max_column + 1)
                if self.is_non_empty(worksheet.cell(row_index, col_index).value)
            )

        return counts

    def _column_non_empty_counts(
        self,
        worksheet: Worksheet,
        row_start: int,
        row_end: int,
    ) -> dict[int, int]:
        """
        Считает количество непустых ячеек в каждой колонке в выбранных строках.

        :param worksheet: Лист Excel для анализа.
        :param row_start: Первая строка рассматриваемой области.
        :param row_end: Последняя строка рассматриваемой области.
        :return: Словарь вида `номер_колонки -> число непустых ячеек`.
        """
        counts: dict[int, int] = {}

        for col_index in range(worksheet.min_column, worksheet.max_column + 1):
            counts[col_index] = sum(
                1
                for row_index in range(row_start, row_end + 1)
                if self.is_non_empty(worksheet.cell(row_index, col_index).value)
            )

        return counts

    @staticmethod
    def _find_bands(
        counts: dict[int, int],
        *,
        min_count: int,
        max_gap: int,
        min_size: int,
    ) -> list[tuple[int, int]]:
        """
        Ищет непрерывные диапазоны строк или колонок, пригодные для таблицы.

        :param counts: Счетчик заполненности строк или колонок.
        :param min_count: Минимальное число заполненных ячеек, чтобы элемент
            считался содержательным.
        :param max_gap: Максимально допустимый разрыв внутри одного диапазона.
        :param min_size: Минимальный размер готового диапазона.
        :return: Список пар `(start, end)` для найденных диапазонов.
        """
        bands: list[tuple[int, int]] = []
        current_start: int | None = None
        current_end: int | None = None
        gap_count = 0

        for index in sorted(counts):
            looks_useful = counts[index] >= min_count

            if looks_useful:
                if current_start is None:
                    current_start = index
                current_end = index
                gap_count = 0
                continue

            if current_start is None:
                continue

            gap_count += 1
            if gap_count > max_gap:
                assert current_end is not None
                if current_end - current_start + 1 >= min_size:
                    bands.append((current_start, current_end))
                current_start = None
                current_end = None
                gap_count = 0

        if current_start is not None and current_end is not None:
            if current_end - current_start + 1 >= min_size:
                bands.append((current_start, current_end))

        return bands

    @staticmethod
    def _region_to_excel_range(region: dict[str, int]) -> str:
        """
        Преобразует координаты области в строковый диапазон Excel.

        :param region: Словарь с границами области.
        :return: Диапазон в нотации Excel, например `C4:F12`.
        """
        start_col = get_column_letter(region["min_col"])
        end_col = get_column_letter(region["max_col"])
        return f"{start_col}{region['min_row']}:{end_col}{region['max_row']}"

    def _region_stats(
        self,
        worksheet: Worksheet,
        region: dict[str, int],
    ) -> dict[str, int | float]:
        """
        Собирает базовую статистику по прямоугольной области листа.

        :param worksheet: Лист Excel.
        :param region: Границы анализируемой области.
        :return: Размеры области, число заполненных ячеек и плотность.
        """
        total_cells = 0
        non_empty_cells = 0

        for row_index in range(region["min_row"], region["max_row"] + 1):
            for col_index in range(region["min_col"], region["max_col"] + 1):
                total_cells += 1
                if self.is_non_empty(worksheet.cell(row_index, col_index).value):
                    non_empty_cells += 1

        rows = region["max_row"] - region["min_row"] + 1
        cols = region["max_col"] - region["min_col"] + 1

        return {
            "rows": rows,
            "cols": cols,
            "total_cells": total_cells,
            "non_empty_cells": non_empty_cells,
            "density": non_empty_cells / total_cells if total_cells else 0.0,
        }

    def _trim_weak_edges(
        self,
        worksheet: Worksheet,
        region: dict[str, int],
        min_filled_ratio: float = 0.3,
    ) -> dict[str, int]:
        """
        Подрезает слабо заполненные края найденного прямоугольника.

        :param worksheet: Лист Excel.
        :param region: Исходная область-кандидат.
        :param min_filled_ratio: Минимальная доля заполнения строки или
            колонки, чтобы сохранить ее в итоговой области.
        :return: Уточненная область без пустых краев.
        """
        min_row = region["min_row"]
        max_row = region["max_row"]
        min_col = region["min_col"]
        max_col = region["max_col"]

        def row_ratio(row_index: int) -> float:
            total = max_col - min_col + 1
            filled = sum(
                1
                for col_index in range(min_col, max_col + 1)
                if self.is_non_empty(worksheet.cell(row_index, col_index).value)
            )
            return filled / total if total else 0.0

        def col_ratio(col_index: int) -> float:
            total = max_row - min_row + 1
            filled = sum(
                1
                for row_index in range(min_row, max_row + 1)
                if self.is_non_empty(worksheet.cell(row_index, col_index).value)
            )
            return filled / total if total else 0.0

        while min_row <= max_row and row_ratio(min_row) < min_filled_ratio:
            min_row += 1
        while max_row >= min_row and row_ratio(max_row) < min_filled_ratio:
            max_row -= 1
        while min_col <= max_col and col_ratio(min_col) < min_filled_ratio:
            min_col += 1
        while max_col >= min_col and col_ratio(max_col) < min_filled_ratio:
            max_col -= 1

        return {
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
        }

    def _row_structure_score(
        self,
        worksheet: Worksheet,
        region: dict[str, int],
    ) -> float:
        """
        Оценивает регулярность строк внутри области.

        :param worksheet: Лист Excel.
        :param region: Анализируемая область.
        :return: Число от `0.0` до `1.0`, где большее значение означает более
            табличную структуру строк.
        """
        counts = [
            sum(
                1
                for col_index in range(region["min_col"], region["max_col"] + 1)
                if self.is_non_empty(worksheet.cell(row_index, col_index).value)
            )
            for row_index in range(region["min_row"], region["max_row"] + 1)
        ]

        if not counts:
            return 0.0

        average = sum(counts) / len(counts)
        if average == 0:
            return 0.0

        deviations = [abs(count - average) / average for count in counts]
        avg_deviation = sum(deviations) / len(deviations)
        return max(0.0, min(1.0, 1.0 - avg_deviation))

    @classmethod
    def _looks_like_number(cls, value: Any) -> bool:
        """
        Проверяет, похоже ли значение на число.

        :param value: Проверяемое значение.
        :return: `True`, если значение можно интерпретировать как число.
        """
        normalized = cls.normalize_cell_value(value)
        if normalized is None:
            return False

        try:
            float(str(normalized).replace(",", ".").replace(" ", ""))
        except ValueError:
            return False

        return True

    def _header_presence_score(
        self,
        worksheet: Worksheet,
        region: dict[str, int],
    ) -> float:
        """
        Оценивает, похожа ли первая строка области на строку заголовков.

        :param worksheet: Лист Excel.
        :param region: Анализируемая область.
        :return: Число от `0.0` до `1.0`, где большее значение означает более
            правдоподобную строку заголовков.
        """
        first_row = region["min_row"]
        second_row = first_row + 1

        if second_row > region["max_row"]:
            return 0.0

        header_values = [
            worksheet.cell(first_row, col_index).value
            for col_index in range(region["min_col"], region["max_col"] + 1)
        ]
        data_values = [
            worksheet.cell(second_row, col_index).value
            for col_index in range(region["min_col"], region["max_col"] + 1)
        ]

        header_values = [value for value in header_values if self.is_non_empty(value)]
        data_values = [value for value in data_values if self.is_non_empty(value)]

        if not header_values or not data_values:
            return 0.0

        header_text_ratio = sum(
            not self._looks_like_number(value)
            for value in header_values
        ) / len(header_values)
        data_number_ratio = sum(
            self._looks_like_number(value)
            for value in data_values
        ) / len(data_values)

        return min(1.0, 0.7 * header_text_ratio + 0.3 * data_number_ratio)

    def _table_likeness_score(
        self,
        worksheet: Worksheet,
        region: dict[str, int],
    ) -> float:
        """
        Вычисляет итоговую оценку похожести области на таблицу.

        :param worksheet: Лист Excel.
        :param region: Анализируемая область.
        :return: Итоговый балл от `0.0` до `1.0`.
        """
        stats = self._region_stats(worksheet, region)
        score = 0.0

        if stats["rows"] >= 3:
            score += 0.2
        elif stats["rows"] >= 2:
            score += 0.1

        if stats["cols"] >= 3:
            score += 0.2
        elif stats["cols"] >= 2:
            score += 0.1

        density = float(stats["density"])
        if density >= 0.5:
            score += 0.25
        elif density >= 0.3:
            score += 0.15
        elif density >= 0.2:
            score += 0.05

        score += 0.25 * self._row_structure_score(worksheet, region)
        score += 0.10 * self._header_presence_score(worksheet, region)

        return round(min(score, 1.0), 4)

    def _is_likely_table_region(
        self,
        worksheet: Worksheet,
        region: dict[str, int],
        *,
        min_density: float = 0.25,
        min_score: float = 0.45,
    ) -> bool:
        """
        Проверяет, можно ли считать область настоящей таблицей.

        :param worksheet: Лист Excel.
        :param region: Анализируемая область.
        :param min_density: Минимальная плотность заполнения.
        :param min_score: Минимальный итоговый балл похожести на таблицу.
        :return: `True`, если область проходит базовые эвристики таблицы.
        """
        stats = self._region_stats(worksheet, region)

        if stats["rows"] < 2 or stats["cols"] < 2:
            return False
        if stats["non_empty_cells"] < 4:
            return False
        if stats["density"] < min_density:
            return False

        return self._table_likeness_score(worksheet, region) >= min_score

    def _build_candidate_regions(self, worksheet: Worksheet) -> list[dict[str, int]]:
        """
        Строит грубые прямоугольники-кандидаты для последующей фильтрации.

        :param worksheet: Лист Excel.
        :return: Список областей-кандидатов без финальной оценки качества.
        """
        row_counts = self._row_non_empty_counts(worksheet)
        row_bands = self._find_bands(
            row_counts,
            min_count=self.min_cells_in_row,
            max_gap=self.max_row_gap,
            min_size=self.min_rows,
        )
        regions: list[dict[str, int]] = []

        for row_start, row_end in row_bands:
            col_counts = self._column_non_empty_counts(
                worksheet,
                row_start=row_start,
                row_end=row_end,
            )
            col_bands = self._find_bands(
                col_counts,
                min_count=self.min_cells_in_col,
                max_gap=self.max_col_gap,
                min_size=self.min_cols,
            )
            for col_start, col_end in col_bands:
                regions.append(
                    {
                        "min_row": row_start,
                        "max_row": row_end,
                        "min_col": col_start,
                        "max_col": col_end,
                    }
                )

        logger.debug(
            "RawWorkbookReader built candidate regions. sheet=%s candidates=%s.",
            worksheet.title,
            len(regions),
        )
        return regions


class XlsxRangeReader:
    """
    Читает конкретный диапазон листа и сохраняет лишние данные для дальнейшего
    разбора.

    Класс нужен в ситуации, когда таблица уже выбрана автоматически или
    пользователем вручную. Он не отбрасывает незнакомые колонки молча, а
    возвращает их как часть диагностики, чтобы следующий слой сам решил,
    что с ними делать.
    """
    def read_range(
        self,
        file_path: str | Path,
        sheet_name: str,
        cell_range: str,
        *,
        entity_type: str | None = None,
    ) -> ExtractedTable:
        """
        Читает выбранный диапазон листа как отдельную таблицу.

        :param file_path: Путь к XLSX-файлу.
        :param sheet_name: Имя листа.
        :param cell_range: Диапазон в нотации Excel, например `C5:N20`.
        :param entity_type: Ожидаемый тип сущности, если пользователь или
            алгоритм уже определили его.
        :return: Извлеченная таблица вместе с диагностикой по заголовкам и
            строкам.
        :raises ValueError: Если лист не найден или диапазон не содержит ни
            строки заголовка, ни строк данных.
        """
        logger.info(
            "XlsxRangeReader range read started. source=%s sheet=%s range=%s entity_type=%s.",
            file_path,
            sheet_name,
            cell_range,
            entity_type,
        )
        workbook = load_workbook(file_path, data_only=True)

        if sheet_name not in workbook.sheetnames:
            logger.error(
                "XlsxRangeReader failed: sheet not found. sheet=%s source=%s.",
                sheet_name,
                file_path,
            )
            raise ValueError(f"Sheet {sheet_name} was not found in workbook.")

        worksheet = workbook[sheet_name]
        extracted_table = self.read_range_from_worksheet(
            worksheet,
            cell_range,
            entity_type=entity_type,
        )
        logger.info(
            "XlsxRangeReader range read finished. sheet=%s range=%s rows=%s errors=%s.",
            extracted_table.sheet,
            extracted_table.range,
            len(extracted_table.rows),
            len(extracted_table.errors),
        )
        return extracted_table

    def read_detected_table(
        self,
        file_path: str | Path,
        table_region: TableRegion,
        *,
        entity_type: str | None = None,
    ) -> ExtractedTable:
        """
        Читает ранее найденную табличную область.

        :param file_path: Путь к XLSX-файлу.
        :param table_region: Найденная область таблицы.
        :param entity_type: Ожидаемый тип сущности.
        :return: Извлеченная таблица для переданного диапазона.
        """
        return self.read_range(
            file_path,
            table_region.sheet,
            table_region.range,
            entity_type=entity_type,
        )

    def read_range_from_worksheet(
        self,
        worksheet: Worksheet,
        cell_range: str,
        *,
        entity_type: str | None = None,
    ) -> ExtractedTable:
        """
        Читает таблицу напрямую из уже открытого листа.

        :param worksheet: Лист Excel.
        :param cell_range: Выбранный диапазон таблицы.
        :param entity_type: Ожидаемый тип сущности.
        :return: Извлеченная таблица вместе с результатами структурной
            проверки.
        :raises ValueError: Если диапазон не содержит ни строки заголовка,
            ни строк данных.
        """
        logger.debug(
            "XlsxRangeReader worksheet range read started. sheet=%s range=%s entity_type=%s.",
            worksheet.title,
            cell_range,
            entity_type,
        )
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        raw_rows = list(
            worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        )

        if not raw_rows:
            raise ValueError(f"Range {cell_range} does not contain rows.")

        headers, warnings = self._build_headers(raw_rows[0])
        if not headers:
            raise ValueError(f"Range {cell_range} does not contain headers.")

        rows = self._build_rows(headers, raw_rows[1:])
        canonical_entity_type = self._normalize_entity_type(entity_type)
        known_headers, unknown_headers, missing_required_headers = (
            self._classify_headers(canonical_entity_type, headers)
        )

        errors: list[str] = []
        if canonical_entity_type and missing_required_headers:
            errors.append(
                "Missing required headers: "
                + ", ".join(missing_required_headers)
                + "."
            )
        if not rows:
            errors.append("Selected range does not contain data rows.")

        if warnings:
            logger.debug(
                "XlsxRangeReader header warnings found. sheet=%s range=%s warnings=%s.",
                worksheet.title,
                cell_range,
                warnings,
            )
        if unknown_headers:
            logger.debug(
                "XlsxRangeReader unknown headers found. sheet=%s range=%s unknown_headers=%s.",
                worksheet.title,
                cell_range,
                unknown_headers,
            )
        if errors:
            logger.warning(
                "XlsxRangeReader range validation failed. sheet=%s range=%s errors=%s.",
                worksheet.title,
                cell_range,
                errors,
            )
        else:
            logger.debug(
                "XlsxRangeReader range validation succeeded. sheet=%s range=%s rows=%s known_headers=%s.",
                worksheet.title,
                cell_range,
                len(rows),
                known_headers,
            )

        return ExtractedTable(
            sheet=worksheet.title,
            range=cell_range,
            entity_type=canonical_entity_type,
            headers=headers,
            rows=rows,
            known_headers=known_headers,
            unknown_headers=unknown_headers,
            missing_required_headers=missing_required_headers,
            warnings=warnings,
            errors=errors,
        )

    def _normalize_entity_type(self, entity_type: str | None) -> str | None:
        """
        Приводит пользовательское имя сущности к каноническому ключу листа.

        :param entity_type: Пользовательское имя сущности или `None`.
        :return: Канонический ключ сущности или `None`.
        """
        if entity_type is None:
            return None

        return normalize_sheet_keys([entity_type])[0]

    def _build_headers(self, raw_headers: tuple[Any, ...]) -> tuple[tuple[str, ...], list[str]]:
        """
        Подготавливает заголовки диапазона и диагностирует проблемы в них.

        :param raw_headers: Сырые значения первой строки диапазона.
        :return: Кортеж итоговых заголовков и список предупреждений.
        """
        headers: list[str] = []
        warnings: list[str] = []
        duplicates: set[str] = set()

        for index, raw_header in enumerate(raw_headers, start=1):
            normalized_header = RawWorkbookReader.normalize_cell_value(raw_header)
            if normalized_header is None:
                generated_header = f"__column_{index}"
                headers.append(generated_header)
                warnings.append(
                    f"Empty header in column {index} was replaced with {generated_header}."
                )
                continue

            header = str(normalized_header).strip()
            if header in headers:
                duplicates.add(header)
                header = f"{header}__{index}"
                warnings.append(
                    f"Duplicate header was renamed to {header}."
                )

            headers.append(header)

        for duplicated_header in sorted(duplicates):
            warnings.append(
                f"Header {duplicated_header} appeared more than once in the selected range."
            )

        return tuple(headers), warnings

    @staticmethod
    def _build_rows(
        headers: tuple[str, ...],
        raw_rows: list[tuple[Any, ...]],
    ) -> list[dict[str, object]]:
        """
        Преобразует строки диапазона в список словарей по заголовкам.

        :param headers: Подготовленные заголовки таблицы.
        :param raw_rows: Сырые строки данных без строки заголовков.
        :return: Непустые строки диапазона в виде словарей.
        """
        rows: list[dict[str, object]] = []

        for raw_row in raw_rows:
            row = {
                header: raw_row[index] if index < len(raw_row) else None
                for index, header in enumerate(headers)
            }
            if all(value is None for value in row.values()):
                continue
            rows.append(row)

        return rows

    def _classify_headers(
        self,
        entity_type: str | None,
        headers: tuple[str, ...],
    ) -> tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
        """
        Разделяет заголовки на известные, неизвестные и обязательные пропущенные.

        :param entity_type: Канонический тип сущности или `None`.
        :param headers: Заголовки выбранного диапазона.
        :return: Кортеж из известных заголовков, неизвестных заголовков и
            пропущенных обязательных заголовков.
        """
        if entity_type is None:
            return (), headers, ()

        classification = classify_tabular_headers(entity_type, headers)
        return (
            classification.known_headers,
            classification.unknown_headers,
            classification.missing_required_headers,
        )


def find_tables_on_sheet(
    worksheet: Worksheet,
    min_score: float = 0.45,
) -> list[TableRegion]:
    """
    Совместимая функциональная точка входа для поиска таблиц на листе.

    :param worksheet: Лист Excel.
    :param min_score: Минимальный итоговый балл похожести на таблицу.
    :return: Список найденных табличных областей.
    """
    return RawWorkbookReader().find_tables_on_sheet(worksheet, min_score=min_score)


def find_tables_in_workbook(
    path: str | Path,
    min_score: float = 0.45,
) -> list[TableRegion]:
    """
    Совместимая функциональная точка входа для поиска таблиц в книге.

    :param path: Путь к XLSX-файлу.
    :param min_score: Минимальный итоговый балл похожести на таблицу.
    :return: Список найденных табличных областей.
    """
    return RawWorkbookReader().find_tables_in_workbook(path, min_score=min_score)


def print_tables(tables: list[TableRegion]) -> None:
    """
    Печатает краткую сводку по найденным таблицам.

    :param tables: Табличные области, найденные в книге.
    :return: `None`.
    """
    if not tables:
        print("Таблицы не найдены.")
        return

    for table in tables:
        print(
            f"{table.sheet} | {table.range} | "
            f"{table.rows}x{table.cols} | "
            f"density={table.density:.2f} | "
            f"score={table.score:.2f}"
        )


def tables_to_dicts(tables: list[TableRegion]) -> list[dict[str, object]]:
    """
    Преобразует найденные области таблиц в список словарей.

    :param tables: Список найденных таблиц.
    :return: Список словарей для сериализации или дальнейшей передачи.
    """
    return [asdict(table) for table in tables]
