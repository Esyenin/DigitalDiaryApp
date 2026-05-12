"""
Высокоуровневый импорт данных приложения из XLSX.

Модуль объединяет три варианта работы с Excel-файлами:

1. Полностью стандартный импорт по внутреннему формату приложения.
2. Чтение произвольного диапазона листа как отдельной таблицы.
3. Подготовка выбранного диапазона к smart- или strict-импорту.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.io_tools.xlsx_importer.data_normalizer import ImportProcessingResult, ImportProcessor
from app.io_tools.xlsx_importer.data_processor import DataProcessingResult, DataProcessor
from app.io_tools.xlsx_config import (
    XLSX_COLUMNS_BY_SHEET,
    XLSX_REQUIRED_COLUMNS_BY_SHEET,
    XLSX_SHEETS_ORDER,
)
from app.io_tools.xlsx_importer.raw_reader import (
    ExtractedTable,
    RawWorkbookReader,
    TableRegion,
    XlsxRangeReader,
)
from app.io_tools.xlsx_importer.strict_import import StrictImportProcessor, StrictImportResult


logger = logging.getLogger(__name__)


class XlsxImporter:
    """
    Координирует чтение XLSX-файлов в разных режимах.

    Класс объединяет низкоуровневое чтение диапазонов и прикладные сценарии
    импорта. Он умеет:

    - импортировать стандартный XLSX формата приложения;
    - искать таблицы в книге без привязки к ячейке `A1`;
    - читать конкретный диапазон как самостоятельную таблицу;
    - запускать smart- или strict-подготовку считанных данных.

    Сам класс не сохраняет данные в базу. Его задача — вернуть структурированный
    результат чтения и предварительной подготовки таблицы.
    """

    def __init__(
        self,
        *,
        raw_workbook_reader: RawWorkbookReader | None = None,
        range_reader: XlsxRangeReader | None = None,
        smart_import_processor: ImportProcessor | None = None,
        data_processor: DataProcessor | None = None,
        strict_import_processor: StrictImportProcessor | None = None,
    ) -> None:
        """
        Создает фасад импорта XLSX.

        :param raw_workbook_reader: Компонент, который ищет кандидаты на
            таблицы в произвольной книге Excel.
        :param range_reader: Компонент, который читает конкретный диапазон
            листа и возвращает диагностическую информацию по заголовкам.
        :param smart_import_processor: Процессор умной подготовки внешних
            таблиц, например со студентами или группами.
        :param data_processor: Слой детальной обработки smart-таблицы с
            сохранением карты распознавания для UI.
        :param strict_import_processor: Процессор строгой подготовки таблиц
            внутреннего формата приложения.
        """
        self.raw_workbook_reader = raw_workbook_reader or RawWorkbookReader()
        self.range_reader = range_reader or XlsxRangeReader()
        self.smart_import_processor = smart_import_processor or ImportProcessor()
        self.data_processor = data_processor or DataProcessor(
            import_processor=self.smart_import_processor,
        )
        self.strict_import_processor = (
            strict_import_processor or StrictImportProcessor()
        )

    def import_data(self, file_path: str | Path) -> dict[str, list[dict[str, object]]]:
        """
        Импортирует XLSX-файл стандартного формата приложения.

        Метод ожидает, что книга уже оформлена в том же формате, который
        использует приложение при стандартном экспорте: известные имена листов,
        ожидаемые заголовки и согласованная структура строк.

        :param file_path: Путь к XLSX-файлу стандартного импорта.
        :return: Словарь вида `имя_листа -> список_строк`, где каждая строка
            представлена словарем значений по заголовкам.
        :raises ValueError: Если книга содержит неизвестные листы или если в
            одном из листов отсутствуют обязательные колонки.
        """
        logger.info("XlsxImporter import started. source=%s.", file_path)
        workbook = load_workbook(file_path)
        workbook_sheet_names = workbook.sheetnames
        logger.debug(
            "XlsxImporter workbook opened. sheets=%s.",
            workbook_sheet_names,
        )
        self._validate_sheet_names(workbook_sheet_names)

        imported_data: dict[str, list[dict[str, object]]] = {}
        ordered_sheet_names = [
            sheet_name
            for sheet_name in XLSX_SHEETS_ORDER
            if sheet_name in workbook_sheet_names
        ]
        ordered_sheet_names.extend(
            sheet_name
            for sheet_name in workbook_sheet_names
            if sheet_name not in ordered_sheet_names
        )

        for sheet_name in ordered_sheet_names:
            worksheet = workbook[sheet_name]
            imported_data[sheet_name] = self._read_sheet(sheet_name, worksheet)

        logger.info(
            "XlsxImporter import finished. source=%s sheets_count=%s.",
            file_path,
            len(imported_data),
        )
        return imported_data

    def find_table_candidates(
        self,
        file_path: str | Path,
        min_score: float = 0.45,
    ) -> list[TableRegion]:
        """
        Ищет в книге области, похожие на таблицы.

        :param file_path: Путь к XLSX-файлу.
        :param min_score: Нижняя граница оценки похожести области на таблицу.
        :return: Список найденных прямоугольных областей с координатами,
            плотностью заполнения и итоговой оценкой.
        """
        logger.info(
            "XlsxImporter raw table detection started. source=%s min_score=%s.",
            file_path,
            min_score,
        )
        tables = self.raw_workbook_reader.find_tables_in_workbook(
            file_path,
            min_score=min_score,
        )
        logger.info(
            "XlsxImporter raw table detection finished. source=%s tables_count=%s.",
            file_path,
            len(tables),
        )
        return tables

    def read_table_range(
        self,
        file_path: str | Path,
        sheet_name: str,
        cell_range: str,
        *,
        entity_type: str | None = None,
    ) -> ExtractedTable:
        """
        Читает произвольный диапазон листа как самостоятельную таблицу.

        :param file_path: Путь к XLSX-файлу.
        :param sheet_name: Имя листа.
        :param cell_range: Диапазон Excel, который нужно прочитать, например
            `C5:N20`.
        :param entity_type: Предполагаемый тип сущности. Нужен, если после
            чтения таблицу надо сопоставить с полями известной модели.
        :return: Объект `ExtractedTable` с заголовками, строками и результатом
            базовой диагностики диапазона.
        """
        logger.info(
            "XlsxImporter range read requested. source=%s sheet=%s range=%s entity_type=%s.",
            file_path,
            sheet_name,
            cell_range,
            entity_type,
        )
        table = self.range_reader.read_range(
            file_path,
            sheet_name,
            cell_range,
            entity_type=entity_type,
        )
        logger.info(
            "XlsxImporter range read finished. sheet=%s range=%s rows=%s warnings=%s errors=%s.",
            table.sheet,
            table.range,
            len(table.rows),
            len(table.warnings),
            len(table.errors),
        )
        return table

    def read_detected_table(
        self,
        file_path: str | Path,
        table_region: TableRegion,
        *,
        entity_type: str | None = None,
    ) -> ExtractedTable:
        """
        Читает ранее найденную табличную область как диапазон Excel.

        :param file_path: Путь к XLSX-файлу.
        :param table_region: Найденная область таблицы.
        :param entity_type: Предполагаемый тип сущности для дальнейшей
            классификации заголовков и строк.
        :return: Извлеченная таблица вместе с диагностикой структуры.
        """
        logger.info(
            "XlsxImporter detected table read requested. source=%s sheet=%s range=%s entity_type=%s.",
            file_path,
            table_region.sheet,
            table_region.range,
            entity_type,
        )
        table = self.range_reader.read_detected_table(
            file_path,
            table_region,
            entity_type=entity_type,
        )
        logger.info(
            "XlsxImporter detected table read finished. sheet=%s range=%s rows=%s warnings=%s errors=%s.",
            table.sheet,
            table.range,
            len(table.rows),
            len(table.warnings),
            len(table.errors),
        )
        return table

    def read_smart_table(
        self,
        file_path: str | Path,
        sheet_name: str,
        cell_range: str,
        *,
        entity_type: str,
    ) -> ImportProcessingResult:
        """
        Читает диапазон и запускает smart-подготовку таблицы.

        :param file_path: Путь к XLSX-файлу.
        :param sheet_name: Имя листа.
        :param cell_range: Диапазон Excel.
        :param entity_type: Тип smart-сущности, например `students` или
            `groups`.
        :return: Результат нормализации, разрешения ссылок и сборки
            `create_payloads` для дальнейшего импорта.
        """
        logger.info(
            "XlsxImporter smart read requested. source=%s sheet=%s range=%s entity_type=%s.",
            file_path,
            sheet_name,
            cell_range,
            entity_type,
        )
        extracted_table = self.read_table_range(
            file_path,
            sheet_name,
            cell_range,
            entity_type=entity_type,
        )
        result = self.smart_import_processor.process_table(extracted_table)
        logger.info(
            "XlsxImporter smart read finished. entity_type=%s payloads=%s errors=%s.",
            result.entity_type,
            len(result.create_payloads),
            len(result.errors),
        )
        return result

    def read_strict_table(
        self,
        file_path: str | Path,
        sheet_name: str,
        cell_range: str,
        *,
        entity_type: str,
    ) -> StrictImportResult:
        """
        Читает диапазон и запускает strict-подготовку таблицы.

        :param file_path: Путь к XLSX-файлу.
        :param sheet_name: Имя листа.
        :param cell_range: Диапазон Excel.
        :param entity_type: Тип strict-сущности внутреннего формата приложения.
        :return: Результат строгой проверки заголовков и валидации строк через
            create-схему нужной сущности.
        """
        logger.info(
            "XlsxImporter strict read requested. source=%s sheet=%s range=%s entity_type=%s.",
            file_path,
            sheet_name,
            cell_range,
            entity_type,
        )
        extracted_table = self.read_table_range(
            file_path,
            sheet_name,
            cell_range,
            entity_type=entity_type,
        )
        result = self.strict_import_processor.process_table(extracted_table)
        logger.info(
            "XlsxImporter strict read finished. entity_type=%s payloads=%s errors=%s.",
            result.entity_type,
            len(result.create_payloads),
            len(result.errors),
        )
        return result

    def process_smart_table(
        self,
        file_path: str | Path,
        sheet_name: str,
        cell_range: str,
        *,
        entity_type: str,
    ) -> DataProcessingResult:
        """
        Читает диапазон и строит полную картину его smart-обработки.

        :param file_path: Путь к XLSX-файлу.
        :param sheet_name: Имя листа.
        :param cell_range: Диапазон Excel.
        :param entity_type: Тип smart-сущности.
        :return: Результат обработки с картой заголовков, построчной
            детализацией и итоговыми payload-ами.
        """
        logger.info(
            "XlsxImporter smart processing requested. source=%s sheet=%s range=%s entity_type=%s.",
            file_path,
            sheet_name,
            cell_range,
            entity_type,
        )
        extracted_table = self.read_table_range(
            file_path,
            sheet_name,
            cell_range,
            entity_type=entity_type,
        )
        result = self.data_processor.process_table(extracted_table)
        logger.info(
            "XlsxImporter smart processing finished. entity_type=%s rows=%s payloads=%s errors=%s.",
            result.entity_type,
            len(result.rows),
            len(result.create_payloads),
            len(result.errors),
        )
        return result

    @staticmethod
    def _validate_sheet_names(sheet_names: list[str]) -> None:
        """
        Проверяет, что стандартная книга не содержит лишних листов.

        :param sheet_names: Имена листов книги.
        :return: `None`.
        :raises ValueError: Если найден хотя бы один лист, который не описан в
            конфигурации стандартного XLSX-формата.
        """
        unknown_sheet_names = [
            sheet_name
            for sheet_name in sheet_names
            if sheet_name not in XLSX_COLUMNS_BY_SHEET
        ]
        if unknown_sheet_names:
            joined = ", ".join(unknown_sheet_names)
            logger.error("XlsxImporter found unknown sheets: %s.", joined)
            raise ValueError(f"Неизвестные листы в XLSX-файле: {joined}.")

    def _read_sheet(
        self,
        sheet_name: str,
        worksheet: Any,
    ) -> list[dict[str, object]]:
        """
        Читает один лист стандартной книги и превращает его строки в словари.

        :param sheet_name: Имя листа.
        :param worksheet: Лист openpyxl.
        :return: Список строк листа, где ключами служат заголовки колонок.
        :raises ValueError: Если в листе отсутствуют обязательные колонки.
        """
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            self._validate_headers(sheet_name, ())
            logger.debug("XlsxImporter read empty sheet=%s.", sheet_name)
            return []

        raw_headers = rows[0]
        headers = tuple(
            str(header).strip()
            for header in raw_headers
            if header is not None and str(header).strip()
        )
        self._validate_headers(sheet_name, headers)
        logger.debug(
            "XlsxImporter validated headers. sheet=%s headers=%s.",
            sheet_name,
            headers,
        )

        imported_rows: list[dict[str, object]] = []
        for raw_row in rows[1:]:
            row_data = {
                header: raw_row[index] if index < len(raw_row) else None
                for index, header in enumerate(headers)
            }
            if self._is_empty_row(row_data):
                continue
            imported_rows.append(row_data)

        logger.debug(
            "XlsxImporter read sheet=%s rows_count=%s headers=%s.",
            sheet_name,
            len(imported_rows),
            headers,
        )
        return imported_rows

    @staticmethod
    def _is_empty_row(row_data: dict[str, object]) -> bool:
        """
        Проверяет, содержит ли строка хоть одно полезное значение.

        :param row_data: Словарь значений строки.
        :return: `True`, если все значения строки отсутствуют, иначе `False`.
        """
        return all(value is None for value in row_data.values())

    @staticmethod
    def _validate_headers(sheet_name: str, headers: tuple[str, ...]) -> None:
        """
        Проверяет обязательные заголовки стандартного листа.

        :param sheet_name: Имя листа.
        :param headers: Колонки, считанные из файла.
        :return: `None`.
        :raises ValueError: Если набор заголовков не содержит хотя бы одну
            колонку, обязательную для данного листа.
        """
        required_headers = XLSX_REQUIRED_COLUMNS_BY_SHEET.get(sheet_name, ())
        missing_headers = [
            header
            for header in required_headers
            if header not in headers
        ]
        if missing_headers:
            joined = ", ".join(missing_headers)
            logger.error(
                "XlsxImporter missing required headers. sheet=%s headers=%s.",
                sheet_name,
                joined,
            )
            raise ValueError(
                f"В листе {sheet_name} отсутствуют обязательные колонки: "
                f"{joined}."
            )
